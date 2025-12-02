from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import check_password_hash, generate_password_hash
from db_connection import create_connection, run_query
from .notifications import get_notifications, mark_notification_read, get_unread_count, create_notification
from .recruitment_change_handler import revert_recruitment_type_change
from extensions import mail
from flask_mail import Message
from datetime import datetime, timedelta
import secrets
import json
import os
import logging
import io
import csv
import requests

logger = logging.getLogger(__name__)


admin_bp = Blueprint("admin", __name__)


@admin_bp.before_request
def restrict_admin_access():
    """
    Checks session before every request.
    - API routes return 401 JSON.
    - Page routes redirect to Login.
    """
    # 1. Allow Login & Static files
    allowed_endpoints = ['admin.login', 'admin.static']
    if request.endpoint in allowed_endpoints:
        return

    # 2. Check Auth
    if 'admin_id' not in session:

        # 3. IDENTIFY API REQUESTS
        # These paths should return JSON error, NOT redirect.
        # (We check prefixes because you have many variations like /approve-reupload/1, /approve-reupload/2...)
        api_prefixes = [
            '/admin/api',
            '/admin/get_job_details',
            '/admin/test-job-report',
            '/admin/job_reports',
            '/admin/applicant_reports',
            '/admin/update_',
            '/admin/approve-',
            '/admin/reject-',
            '/admin/delete-',
            '/admin/reupload-',
        ]

        is_api_call = (
            request.is_json or
            any(request.path.startswith(prefix) for prefix in api_prefixes)
        )

        if is_api_call:
            return jsonify({"success": False, "message": "Unauthorized: Session expired."}), 401

        # 4. Standard Pages -> Redirect to Login
        flash("Please log in to access the admin area.", "danger")
        return redirect(url_for('admin.login'))


def _to_int(value):
    """Convert DB numeric values (Decimal, None) to plain int safely."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _parse_multi(args, key):
    val = args.get(key)
    if not val:
        return []
    # CHANGE THIS: Split by pipe '|' instead of comma ','
    return [x.strip() for x in val.split('|') if x.strip()]


def _matches_age_bracket(age, age_brackets):
    """Check if age matches any of the specified age brackets."""
    # If no age brackets are selected, include ALL ages
    if not age_brackets:
        return True

    if not age:
        return False

    age_int = int(age)

    for bracket in age_brackets:
        if bracket == "60+":
            if age_int >= 60:
                return True
        elif '-' in bracket:
            try:
                start, end = bracket.split('-')
                start_age = int(start)
                end_age = int(end)
                if start_age <= age_int <= end_age:
                    return True
            except ValueError:
                continue

    return False


def build_applicants_filters(args, alias="a"):
    """Build WHERE clause + params for applicant analytics filters."""
    clauses = ["1=1"]
    params = []

    # Time period (created_at) - no changes needed
    date_from = args.get("date_from")
    date_to = args.get("date_to")
    quick_range = args.get("quick_range")

    if quick_range and not (date_from or date_to):
        today = datetime.today().date()
        if quick_range == "last_30":
            date_from = (today - timedelta(days=30)).isoformat()
            date_to = today.isoformat()
        elif quick_range == "ytd":
            date_from = f"{today.year}-01-01"
            date_to = today.isoformat()
        elif quick_range == "qtd":
            quarter = (today.month - 1) // 3 + 1
            start_month = 3 * (quarter - 1) + 1
            date_from = f"{today.year}-{start_month:02d}-01"
            date_to = today.isoformat()

    if date_from:
        clauses.append(f"{alias}.created_at >= %s")
        params.append(date_from)
    if date_to:
        clauses.append(f"{alias}.created_at < DATE_ADD(%s, INTERVAL 1 DAY)")
        params.append(date_to)

    # Status / is_active - no changes needed (these are usually standardized)
    statuses = _parse_multi(args, "applicant_status")
    if statuses:
        placeholders = ",".join(["%s"] * len(statuses))
        clauses.append(f"{alias}.status IN ({placeholders})")
        params.extend(statuses)

    is_active_vals = _parse_multi(args, "applicant_is_active")
    if is_active_vals:
        placeholders = ",".join(["%s"] * len(is_active_vals))
        clauses.append(f"{alias}.is_active IN ({placeholders})")
        params.extend(is_active_vals)

    # Demographics - MAKE CASE-INSENSITIVE
    sexes = _parse_multi(args, "sex")
    if sexes:
        placeholders = ",".join(["%s"] * len(sexes))
        clauses.append(f"UPPER({alias}.sex) IN ({placeholders})")
        params.extend([sex.upper() for sex in sexes])

    educations = _parse_multi(args, "education")
    if educations:
        placeholders = ",".join(["%s"] * len(educations))
        clauses.append(f"UPPER({alias}.education) IN ({placeholders})")
        params.extend([education.upper() for education in educations])

    # Boolean fields - no changes needed
    is_pwd_vals = _parse_multi(args, "is_pwd")
    if is_pwd_vals:
        placeholders = ",".join(["%s"] * len(is_pwd_vals))
        clauses.append(f"{alias}.is_pwd IN ({placeholders})")
        params.extend(is_pwd_vals)

    has_work_exp_vals = _parse_multi(args, "has_work_exp")
    if has_work_exp_vals:
        placeholders = ",".join(["%s"] * len(has_work_exp_vals))
        clauses.append(f"{alias}.has_work_exp IN ({placeholders})")
        params.extend(has_work_exp_vals)

    # Location - MAKE CASE-INSENSITIVE
    provinces = _parse_multi(args, "applicant_province")
    if provinces:
        placeholders = ",".join(["%s"] * len(provinces))
        clauses.append(f"UPPER({alias}.province) IN ({placeholders})")
        params.extend([province.upper() for province in provinces])

    cities = _parse_multi(args, "applicant_city")
    if cities:
        placeholders = ",".join(["%s"] * len(cities))
        clauses.append(f"UPPER({alias}.city) IN ({placeholders})")
        params.extend([city.upper() for city in cities])

    barangays = _parse_multi(args, "applicant_barangay")
    if barangays:
        placeholders = ",".join(["%s"] * len(barangays))
        clauses.append(f"UPPER({alias}.barangay) IN ({placeholders})")
        params.extend([barangay.upper() for barangay in barangays])

    # Age bracket filtering - parse the parameter so it's available for post-processing
    age_brackets = _parse_multi(args, "age_bracket")
    # Note: Age bracket filtering is applied in Python after SQL query
    # We just need to ensure the parameter is parsed

    return " AND ".join(clauses), tuple(params)


def build_employers_filters(args, alias="e"):
    """Build WHERE clause + params for employer analytics filters."""
    clauses = ["1=1"]
    params = []

    # Time period (created_at) - no changes needed here
    date_from = args.get("date_from")
    date_to = args.get("date_to")
    quick_range = args.get("quick_range")

    if quick_range and not (date_from or date_to):
        today = datetime.today().date()
        if quick_range == "last_30":
            date_from = (today - timedelta(days=30)).isoformat()
            date_to = today.isoformat()
        elif quick_range == "ytd":
            date_from = f"{today.year}-01-01"
            date_to = today.isoformat()
        elif quick_range == "qtd":
            quarter = (today.month - 1) // 3 + 1
            start_month = 3 * (quarter - 1) + 1
            date_from = f"{today.year}-{start_month:02d}-01"
            date_to = today.isoformat()

    if date_from:
        clauses.append(f"{alias}.created_at >= %s")
        params.append(date_from)
    if date_to:
        clauses.append(f"{alias}.created_at < DATE_ADD(%s, INTERVAL 1 DAY)")
        params.append(date_to)

    # Status / is_active - no changes needed (these are usually standardized)
    statuses = _parse_multi(args, "employer_status")
    if statuses:
        placeholders = ",".join(["%s"] * len(statuses))
        clauses.append(f"{alias}.status IN ({placeholders})")
        params.extend(statuses)

    is_active_vals = _parse_multi(args, "employer_is_active")
    if is_active_vals:
        placeholders = ",".join(["%s"] * len(is_active_vals))
        clauses.append(f"{alias}.is_active IN ({placeholders})")
        params.extend(is_active_vals)

    # Industry - MAKE CASE-INSENSITIVE
    industries = _parse_multi(args, "industry")
    print(f"DEBUG - Industries filter: {industries}")  # ADD THIS LINE

    if industries:
        placeholders = ",".join(["%s"] * len(industries))
        clauses.append(f"UPPER({alias}.industry) IN ({placeholders})")
        params.extend([industry.upper() for industry in industries])

    # Recruitment type - MAKE CASE-INSENSITIVE
    rec_types = _parse_multi(args, "recruitment_type")
    print(f"DEBUG - Recruitment types filter: {rec_types}")  # ADD THIS LINE

    if rec_types:
        placeholders = ",".join(["%s"] * len(rec_types))
        clauses.append(f"UPPER({alias}.recruitment_type) IN ({placeholders})")
        params.extend([rec_type.upper() for rec_type in rec_types])

    # Location
    provinces = _parse_multi(args, "employer_province")
    if provinces:
        placeholders = ",".join(["%s"] * len(provinces))
        # Simply check if the province column matches ANY of the selected provinces
        clauses.append(f"UPPER({alias}.province) IN ({placeholders})")
        params.extend([province.upper() for province in provinces])

    cities = _parse_multi(args, "employer_city")
    barangays = _parse_multi(args, "employer_barangay")

    print(f"DEBUG - Provinces filter: {provinces}")
    print(f"DEBUG - Cities filter: {cities}")

    if provinces:
        placeholders = ",".join(["%s"] * len(provinces))
        manila_indicators = {"MANILA", "METRO MANILA",
                             "NCR", "NATIONAL CAPITAL REGION"}
        has_manila = any(p.upper() in manila_indicators for p in provinces)

        if has_manila:
            print("DEBUG - Manila special case triggered")
            clauses.append(f"UPPER({alias}.city) = 'MANILA'")
        else:
            clauses.append(f"UPPER({alias}.province) IN ({placeholders})")
            params.extend([province.upper() for province in provinces])

    if cities:
        placeholders = ",".join(["%s"] * len(cities))
        clauses.append(f"UPPER({alias}.city) IN ({placeholders})")
        params.extend([city.upper() for city in cities])

    if barangays:
        placeholders = ",".join(["%s"] * len(barangays))
        clauses.append(f"UPPER({alias}.barangay) IN ({placeholders})")
        params.extend([barangay.upper() for barangay in barangays])

    print(f"DEBUG - Final WHERE clauses: {clauses}")
    print(f"DEBUG - Final params: {params}")

    return " AND ".join(clauses), tuple(params)


def build_jobs_filters(args, alias="j"):
    """Build WHERE clause + params for job analytics filters."""
    clauses = ["1=1"]
    params = []

    date_from = args.get("date_from")
    date_to = args.get("date_to")
    quick_range = args.get("quick_range")

    if quick_range and not (date_from or date_to):
        today = datetime.today().date()
        if quick_range == "last_30":
            date_from = (today - timedelta(days=30)).isoformat()
            date_to = today.isoformat()
        elif quick_range == "ytd":
            date_from = f"{today.year}-01-01"
            date_to = today.isoformat()
        elif quick_range == "qtd":
            quarter = (today.month - 1) // 3 + 1
            start_month = 3 * (quarter - 1) + 1
            date_from = f"{today.year}-{start_month:02d}-01"
            date_to = today.isoformat()

    if date_from:
        clauses.append(f"{alias}.created_at >= %s")
        params.append(date_from)
    if date_to:
        clauses.append(f"{alias}.created_at < DATE_ADD(%s, INTERVAL 1 DAY)")
        params.append(date_to)

    job_statuses = _parse_multi(args, "job_status")
    if job_statuses:
        # Map filter values to database values
        status_mapping = {
            "Active": "active",
            "Inactive": "inactive",
            "Archived": "archived",
            "Suspended": "suspended"
        }
        db_statuses = [status_mapping.get(status, status)
                       for status in job_statuses]

        placeholders = ",".join(["%s"] * len(db_statuses))
        clauses.append(f"{alias}.status IN ({placeholders})")
        params.extend(db_statuses)

    work_schedules = _parse_multi(args, "work_schedule")
    if work_schedules:
        # Map user-friendly filter values to actual database values
        schedule_mapping = {
            "Full-Time": "full-time",
            "Part-Time": "part-time",
            "Contract": "contract",
            "Freelance": "freelance"
        }
        db_schedules = [schedule_mapping.get(
            schedule, schedule) for schedule in work_schedules]

        placeholders = ",".join(["%s"] * len(db_schedules))
        clauses.append(f"UPPER({alias}.work_schedule) IN ({placeholders})")
        params.extend([schedule.upper() for schedule in db_schedules])

    return " AND ".join(clauses), tuple(params)


def build_applications_filters(args, alias="a"):
    """Build WHERE clause + params for application analytics filters."""
    clauses = ["1=1"]
    params = []

    date_from = args.get("date_from")
    date_to = args.get("date_to")
    quick_range = args.get("quick_range")

    if quick_range and not (date_from or date_to):
        today = datetime.today().date()
        if quick_range == "last_30":
            date_from = (today - timedelta(days=30)).isoformat()
            date_to = today.isoformat()
        elif quick_range == "ytd":
            date_from = f"{today.year}-01-01"
            date_to = today.isoformat()
        elif quick_range == "qtd":
            quarter = (today.month - 1) // 3 + 1
            start_month = 3 * (quarter - 1) + 1
            date_from = f"{today.year}-{start_month:02d}-01"
            date_to = today.isoformat()

    # Use applied_at for applications
    if date_from:
        clauses.append(f"{alias}.applied_at >= %s")
        params.append(date_from)
    if date_to:
        clauses.append(f"{alias}.applied_at < DATE_ADD(%s, INTERVAL 1 DAY)")
        params.append(date_to)

    statuses = _parse_multi(args, "application_status")
    if statuses:
        # Application status values - these should match your modal exactly
        # No mapping needed since filter values match database values
        placeholders = ",".join(["%s"] * len(statuses))
        clauses.append(f"UPPER({alias}.status) IN ({placeholders})")
        params.extend([status.upper() for status in statuses])

    # Filter by work_schedule via join alias "j"
    work_schedules = _parse_multi(args, "work_schedule")
    if work_schedules:
        # Map user-friendly filter values to actual database values
        schedule_mapping = {
            "Full-Time": "full-time",
            "Part-Time": "part-time",
            "Contract": "contract",
            "Freelance": "freelance"
        }
        db_schedules = [schedule_mapping.get(
            schedule, schedule) for schedule in work_schedules]

        placeholders = ",".join(["%s"] * len(db_schedules))
        clauses.append(f"UPPER(j.work_schedule) IN ({placeholders})")
        params.extend([schedule.upper() for schedule in db_schedules])

    return " AND ".join(clauses), tuple(params)


@admin_bp.route("/api/analytics/export", methods=["POST", "GET"])
def analytics_export():
    """Export filtered data for a module as CSV, XLSX, or PDF."""

    # Handle both POST (JSON) and GET (query params) for PDF
    if request.method == "GET":
        module = request.args.get("module")
        export_format = request.args.get("format", "csv").lower()
        filters = {}
        # Parse filters from query parameters
        for key in request.args:
            if key not in ['module', 'format']:
                value = request.args.get(key)
                if value and ',' in value:
                    filters[key] = value.split(',')
                else:
                    filters[key] = value
    else:
        payload = request.get_json(silent=True) or {}
        module = payload.get("module")
        export_format = (payload.get("format") or "csv").lower()
        filters = payload.get("filters") or {}

    # Input validation with JSON responses
    if not module:
        return jsonify({"success": False, "message": "Module parameter is required"}), 400

    if module not in {"applicants", "employers", "jobs_applications"}:
        return jsonify({"success": False, "message": "Invalid module"}), 400

    if export_format not in {"csv", "xlsx", "pdf"}:
        return jsonify({"success": False, "message": "Format not supported"}), 400

    # Build args-like object for helpers
    class _Args:
        def __init__(self, d):
            self._d = d

        def get(self, key, default=None):
            v = self._d.get(key, default)
            if isinstance(v, list):
                return ",".join(v)
            return v

    args_obj = _Args(filters)

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        if module == "applicants":
            where_sql, params = build_applicants_filters(args_obj, alias="a")
            rows = run_query(
                conn,
                f"""
                SELECT
                  a.applicant_id,
                  a.applicant_code,
                  a.first_name,
                  a.last_name,
                  a.middle_name,
                  a.age,
                  a.sex,
                  a.phone,
                  a.email,
                  a.is_from_lipa,
                  a.province,
                  a.city,
                  a.barangay,
                  a.education,
                  a.is_pwd,
                  a.pwd_type,
                  a.has_work_exp as has_work_exp,
                  a.years_experience,
                  a.registration_reason,
                  a.status,
                  a.is_active,
                  DATE(a.created_at) as created_date
                FROM applicants a
                WHERE {where_sql}
                ORDER BY a.created_at DESC
                """,
                params,
                fetch="all",
            ) or []

            # Apply age bracket filtering in Python since it's complex for SQL
            age_brackets = _parse_multi(args_obj, "age_bracket")
            if age_brackets:
                rows = [row for row in rows if _matches_age_bracket(
                    row['age'], age_brackets)]

        elif module == "employers":
            where_sql, params = build_employers_filters(args_obj, alias="e")
            rows = run_query(
                conn,
                f"""
                SELECT
                  e.employer_id,
                  e.employer_code,
                  e.employer_name,
                  e.industry,
                  e.recruitment_type,
                  e.contact_person,
                  e.phone,
                  e.email,
                  e.province,
                  e.city,
                  e.barangay,
                  e.status,
                  e.is_active,
                  DATE(e.created_at) as created_date
                FROM employers e
                WHERE {where_sql}
                ORDER BY e.created_at DESC
                """,
                params,
                fetch="all",
            ) or []
        else:  # jobs_applications
            where_sql, params = build_applications_filters(
                args_obj, alias="app")
            rows = run_query(
                conn,
                f"""
                SELECT
                  app.id AS application_id,
                  app.applicant_id,
                  app.job_id,
                  app.status AS application_status,
                  DATE(app.applied_at) AS application_date,
                  j.job_position,
                  j.work_schedule,
                  j.status AS job_status,
                  a.first_name,
                  a.last_name,
                  e.employer_name
                FROM applications app
                LEFT JOIN jobs j ON app.job_id = j.job_id
                LEFT JOIN applicants a ON app.applicant_id = a.applicant_id
                LEFT JOIN employers e ON j.employer_id = e.employer_id
                WHERE {where_sql}
                ORDER BY app.applied_at DESC
                """,
                params,
                fetch="all",
            ) or []

        if not rows:
            return jsonify({"success": False, "message": "No records match the current filters."}), 400

        fieldnames = list(rows[0].keys())

        if export_format == "csv":
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            for r in rows:
                writer.writerow(r)
            mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
            filename = f"{module}_export.csv"

            return send_file(
                mem,
                mimetype="text/csv",
                as_attachment=True,
                download_name=filename,
            )
        elif export_format == "xlsx":
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
                from openpyxl.utils import get_column_letter
            except ImportError:
                return jsonify({"success": False, "message": "openpyxl is required for XLSX export"}), 500

            wb = Workbook()
            ws = wb.active

            # Add headers
            ws.append(fieldnames)

            # Make header row bold
            for col in range(1, len(fieldnames) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)

            # Add data rows
            for r in rows:
                ws.append([r.get(f) for f in fieldnames])

            # Auto-adjust column widths
            for column_cells in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    try:
                        if cell.value:
                            # Calculate length of cell content
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                # Set column width with some padding (max 50 characters wide)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            mem = io.BytesIO()
            wb.save(mem)
            mem.seek(0)

            filename = f"{module}_export.xlsx"

            return send_file(
                mem,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )
        else:  # pdf
            try:
                from reportlab.lib.pagesizes import letter, landscape
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib import colors
                from reportlab.lib.units import inch
            except ImportError:
                return jsonify({"success": False, "message": "reportlab is required for PDF export"}), 500

            mem = io.BytesIO()

            # Use landscape for better table viewing
            doc = SimpleDocTemplate(mem, pagesize=landscape(letter))
            elements = []

            # Add title
            styles = getSampleStyleSheet()
            title = Paragraph(
                f"{module.replace('_', ' ').title()} Export", styles['Title'])
            elements.append(title)

            # Add export timestamp
            timestamp = Paragraph(
                f"Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
            elements.append(timestamp)
            elements.append(Spacer(1, 0.2 * inch))

            # Prepare data for table - limit columns for PDF readability
            max_columns = 8  # Limit columns for PDF readability
            if len(fieldnames) > max_columns:
                # Select most important columns for PDF
                important_fields = []
                if module == "applicants":
                    important_fields = ['applicant_code', 'first_name',
                                        'last_name', 'age', 'sex', 'city', 'status', 'created_date']
                elif module == "employers":
                    important_fields = ['employer_code', 'employer_name', 'industry',
                                        'recruitment_type', 'city', 'status', 'created_date']
                else:  # jobs_applications
                    important_fields = ['application_id', 'first_name', 'last_name',
                                        'job_position', 'application_status', 'employer_name', 'application_date']

                # Use important fields that actually exist in the data
                fieldnames_display = [
                    f for f in important_fields if f in fieldnames]
                # Add any missing important fields from original data
                for f in fieldnames:
                    if f not in fieldnames_display and len(fieldnames_display) < max_columns:
                        fieldnames_display.append(f)
            else:
                fieldnames_display = fieldnames

            data = [fieldnames_display]  # Header row

            for row in rows:
                data.append([str(row.get(field, ''))
                            for field in fieldnames_display])

            # Create table
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1b5e20')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [colors.white, colors.lightgrey])
            ]))
            elements.append(table)

            doc.build(elements)
            mem.seek(0)

            filename = f"{module}_export.pdf"

            return send_file(
                mem,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=filename,
            )

    except Exception as exc:
        print("[analytics] export error:", exc)
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Failed to export data: {str(exc)}"}), 500
    finally:
        conn.close()


# ===== Admin Home (Dashboard with notifications) =====
@admin_bp.route("/home")
def admin_home():
    return render_template("Admin/admin_home.html")


@admin_bp.route("/api/analytics/summary", methods=["GET"])
def admin_analytics_summary():

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        applicants_row = run_query(
            conn, "SELECT COUNT(*) AS total FROM applicants", fetch="one") or {}
        employers_row = run_query(
            conn, "SELECT COUNT(*) AS total FROM employers", fetch="one") or {}
        jobs_row = run_query(
            conn, "SELECT COUNT(*) AS total FROM jobs WHERE status = 'active'", fetch="one") or {}
        applications_row = run_query(
            conn, "SELECT COUNT(*) AS total FROM applications", fetch="one") or {}

        payload = {
            "totalApplicants": _to_int(applicants_row.get("total")),
            "totalEmployers": _to_int(employers_row.get("total")),
            "activeJobs": _to_int(jobs_row.get("total")),
            "totalApplications": _to_int(applications_row.get("total")),
        }
        return jsonify({"success": True, "data": payload})
    except Exception as exc:
        print("[analytics] summary error:", exc)
        return jsonify({"success": False, "message": "Failed to load summary"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/applicants-per-month", methods=["GET"])
def admin_analytics_applicants_per_month():

    year_filter = request.args.get("year", type=int)

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_applicants_filters(request.args, alias="a")
        if year_filter:
            where_sql += " AND YEAR(a.created_at) = %s"
            params = (*params, year_filter)

        # Get all applicants with their creation month and age
        rows = run_query(
            conn,
            f"""
            SELECT
                DATE_FORMAT(a.created_at, '%Y-%m') AS month_key,
                DATE_FORMAT(a.created_at, '%b %Y') AS label,
                MONTH(a.created_at) AS month_num,
                YEAR(a.created_at) AS year_num,
                a.age
            FROM applicants a
            WHERE {where_sql}
            ORDER BY year_num ASC, month_num ASC
            """,
            params,
            fetch="all",
        ) or []

        # Apply age bracket filtering in Python
        age_brackets = _parse_multi(request.args, "age_bracket")
        if age_brackets:
            rows = [row for row in rows if _matches_age_bracket(
                row['age'], age_brackets)]

        # Group by month after filtering
        monthly_counts = {}
        for row in rows:
            month_key = row.get("month_key")
            if month_key not in monthly_counts:
                monthly_counts[month_key] = {
                    "month": _to_int(row.get("month_num")),
                    "year": _to_int(row.get("year_num")),
                    "label": row.get("label"),
                    "count": 0
                }
            monthly_counts[month_key]["count"] += 1

        data = list(monthly_counts.values())
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] applicants-per-month error:", exc)
        return jsonify({"success": False, "message": "Failed to load applicants per month"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/applications-by-category", methods=["GET"])
def admin_analytics_applications_by_category():

    month_filter = request.args.get("month", type=int)
    year_filter = request.args.get("year", type=int)
    category_filter = request.args.get("category", type=str)

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        query = """
            SELECT
                COALESCE(NULLIF(j.work_schedule, ''), 'Unspecified') AS category,
                COUNT(a.id) AS total
            FROM applications a
            JOIN jobs j ON a.job_id = j.job_id
            WHERE 1=1
        """
        params = []

        if month_filter:
            query += " AND MONTH(a.created_at) = %s"
            params.append(month_filter)
        if year_filter:
            query += " AND YEAR(a.created_at) = %s"
            params.append(year_filter)
        if category_filter:
            query += " AND j.work_schedule = %s"
            params.append(category_filter)

        query += " GROUP BY category ORDER BY total DESC"

        rows = run_query(conn, query, tuple(params)
                         if params else None, fetch="all") or []

        data = [
            {"category": row.get("category"),
             "count": _to_int(row.get("total"))}
            for row in rows
        ]
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] applications-by-category error:", exc)
        return jsonify({"success": False, "message": "Failed to load applications by category"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/hiring-ratio", methods=["GET"])
def admin_analytics_hiring_ratio():

    month_filter = request.args.get("month", type=int)
    year_filter = request.args.get("year", type=int)

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        query = """
            SELECT COALESCE(status, 'Pending') AS status, COUNT(*) AS total
            FROM applications
            WHERE 1=1
        """
        params = []

        if month_filter:
            query += " AND MONTH(created_at) = %s"
            params.append(month_filter)
        if year_filter:
            query += " AND YEAR(created_at) = %s"
            params.append(year_filter)

        query += " GROUP BY status"

        rows = run_query(conn, query, tuple(params)
                         if params else None, fetch="all") or []

        breakdown = {
            row.get("status"): _to_int(row.get("total"))
            for row in rows
        }
        hired = breakdown.get("Hired", 0)
        total = sum(breakdown.values())
        not_hired = max(total - hired, 0)

        data = {
            "hired": hired,
            "not_hired": not_hired,
            "breakdown": breakdown,
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] hiring-ratio error:", exc)
        return jsonify({"success": False, "message": "Failed to load hiring ratio"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/filters/applicants/locations", methods=["GET"])
def applicants_location_filters():

    level = request.args.get("level", "province").lower()
    parent = request.args.get("parent")

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        # For filter dropdowns, we want ALL available locations, not filtered ones
        if level == "province":
            rows = run_query(
                conn,
                """
                SELECT DISTINCT UPPER(province) AS value
                FROM applicants
                WHERE province IS NOT NULL AND province <> ''
                ORDER BY value ASC
                """,
                fetch="all",
            ) or []
        elif level == "city":
            if not parent:
                return jsonify({"success": True, "data": []})
            rows = run_query(
                conn,
                """
                SELECT DISTINCT UPPER(city) AS value
                FROM applicants
                WHERE UPPER(province) = UPPER(%s) AND city IS NOT NULL AND city <> ''
                ORDER BY value ASC
                """,
                (parent,),
                fetch="all",
            ) or []
        elif level == "barangay":
            if not parent:
                return jsonify({"success": True, "data": []})
            rows = run_query(
                conn,
                """
                SELECT DISTINCT UPPER(barangay) AS value
                FROM applicants
                WHERE UPPER(city) = UPPER(%s) AND barangay IS NOT NULL AND barangay <> ''
                ORDER BY value ASC
                """,
                (parent,),
                fetch="all",
            ) or []
        else:
            return jsonify({"success": False, "message": "Invalid level"}), 400

        values = [row.get("value") for row in rows if row.get("value")]
        return jsonify({"success": True, "data": values})
    except Exception as exc:
        print("[filters] applicants locations error:", exc)
        return jsonify({"success": False, "message": "Failed to load locations"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/filters/employers/locations", methods=["GET"])
def employers_location_filters():

    level = request.args.get("level", "province").lower()
    parent = request.args.get("parent")

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        # For filter dropdowns, we want ALL available locations, not filtered ones
        if level == "province":
            rows = run_query(
                conn,
                """
                SELECT DISTINCT UPPER(province) AS value
                FROM employers
                WHERE province IS NOT NULL AND province <> ''
                ORDER BY value ASC
                """,
                fetch="all",
            ) or []
        elif level == "city":
            if not parent:
                return jsonify({"success": True, "data": []})
            rows = run_query(
                conn,
                """
                SELECT DISTINCT UPPER(city) AS value
                FROM employers
                WHERE UPPER(province) = UPPER(%s) AND city IS NOT NULL AND city <> ''
                ORDER BY value ASC
                """,
                (parent,),
                fetch="all",
            ) or []
        elif level == "barangay":
            if not parent:
                return jsonify({"success": True, "data": []})
            rows = run_query(
                conn,
                """
                SELECT DISTINCT UPPER(barangay) AS value
                FROM employers
                WHERE UPPER(city) = UPPER(%s) AND barangay IS NOT NULL AND barangay <> ''
                ORDER BY value ASC
                """,
                (parent,),
                fetch="all",
            ) or []
        else:
            return jsonify({"success": False, "message": "Invalid level"}), 400

        values = [row.get("value") for row in rows if row.get("value")]
        return jsonify({"success": True, "data": values})
    except Exception as exc:
        print("[filters] employers locations error:", exc)
        return jsonify({"success": False, "message": "Failed to load locations"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/applicants-by-province", methods=["GET"])
def admin_analytics_applicants_by_province():

    month_filter = request.args.get("month", type=int)
    year_filter = request.args.get("year", type=int)

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_applicants_filters(request.args, alias="a")

        # Add month/year filters if provided
        if month_filter:
            where_sql += " AND MONTH(a.created_at) = %s"
            params = (*params, month_filter)
        if year_filter:
            where_sql += " AND YEAR(a.created_at) = %s"
            params = (*params, year_filter)

        # Get all applicants with province and age
        rows = run_query(
            conn,
            f"""
            SELECT
                COALESCE(NULLIF(a.province, ''), 'Unspecified') AS province,
                a.age
            FROM applicants a
            WHERE {where_sql}
            """,
            params,
            fetch="all",
        ) or []

        # Apply age bracket filtering in Python
        age_brackets = _parse_multi(request.args, "age_bracket")
        if age_brackets:
            rows = [row for row in rows if _matches_age_bracket(
                row['age'], age_brackets)]

        # Count by province after filtering
        province_counts = {}
        for row in rows:
            province = row.get("province")
            province_counts[province] = province_counts.get(province, 0) + 1

        data = [
            {"province": province, "count": count}
            for province, count in province_counts.items()
        ]
        data.sort(key=lambda x: x["count"], reverse=True)

        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] applicants-by-province error:", exc)
        return jsonify({"success": False, "message": "Failed to load applicant locations"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/employers-by-industry", methods=["GET"])
def admin_analytics_employers_by_industry():

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        rows = run_query(
            conn,
            """
            SELECT
                COALESCE(NULLIF(industry, ''), 'Unspecified') AS industry,
                COUNT(*) AS total
            FROM employers
            GROUP BY industry
            ORDER BY total DESC
            """,
            fetch="all",
        ) or []

        data = [
            {"industry": row.get("industry"),
             "count": _to_int(row.get("total"))}
            for row in rows
        ]
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] employers-by-industry error:", exc)
        return jsonify({"success": False, "message": "Failed to load employer industries"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/applicants/summary", methods=["GET"])
def applicants_summary():
    """Applicants volume & active count, respecting filters."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_applicants_filters(request.args, alias="a")

        # Get all applicants first
        rows = run_query(
            conn,
            f"""
            SELECT
              a.applicant_id,
              a.age
            FROM applicants a
            WHERE {where_sql}
            """,
            params,
            fetch="all",
        ) or []

        # Apply age bracket filtering in Python
        age_brackets = _parse_multi(request.args, "age_bracket")
        if age_brackets:
            rows = [row for row in rows if _matches_age_bracket(
                row['age'], age_brackets)]

        # Now calculate summary based on filtered rows
        total_registered = len(rows)
        active_count = run_query(
            conn,
            f"""
            SELECT COUNT(*) AS active_count
            FROM applicants a
            WHERE {where_sql} AND a.is_active = 1
            """,
            params,
            fetch="one",
        ) or {}

        # Apply age bracket filtering to active count too
        if age_brackets:
            active_rows = run_query(
                conn,
                f"""
                SELECT a.applicant_id, a.age
                FROM applicants a
                WHERE {where_sql} AND a.is_active = 1
                """,
                params,
                fetch="all",
            ) or []
            active_rows = [row for row in active_rows if _matches_age_bracket(
                row['age'], age_brackets)]
            active_count = {"active_count": len(active_rows)}
        else:
            active_count = {"active_count": _to_int(
                active_count.get("active_count"))}

        data = {
            "total_registered": total_registered,
            "active_applicants": active_count["active_count"],
            "new_registrations": total_registered,  # This might need adjustment
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] applicants_summary error:", exc)
        return jsonify({"success": False, "message": "Failed to load applicants summary"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/applicants/demographics", methods=["GET"])
def applicants_demographics():
    """Applicants by sex, education, and age groups."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_applicants_filters(request.args, alias="a")

        # Get all applicants first
        all_applicants = run_query(
            conn,
            f"""
            SELECT
              a.applicant_id,
              a.sex,
              a.education,
              a.age
            FROM applicants a
            WHERE {where_sql}
            """,
            params,
            fetch="all",
        ) or []

        # Apply age bracket filtering in Python
        age_brackets = _parse_multi(request.args, "age_bracket")
        if age_brackets:
            all_applicants = [row for row in all_applicants if _matches_age_bracket(
                row['age'], age_brackets)]

        # Calculate demographics from filtered data
        sex_counts = {}
        education_counts = {}
        age_group_counts = {
            "Under 18": 0,
            "18-24": 0,
            "25-34": 0,
            "35-44": 0,
            "45+": 0,
            "Unspecified": 0
        }

        for applicant in all_applicants:
            # Sex distribution
            sex = applicant.get("sex") or "Unspecified"
            sex_counts[sex] = sex_counts.get(sex, 0) + 1

            # Education distribution
            education = applicant.get("education") or "Unspecified"
            education_counts[education] = education_counts.get(
                education, 0) + 1

            # Age groups
            age = applicant.get("age")
            if not age:
                age_group_counts["Unspecified"] += 1
            elif age < 18:
                age_group_counts["Under 18"] += 1
            elif 18 <= age <= 24:
                age_group_counts["18-24"] += 1
            elif 25 <= age <= 34:
                age_group_counts["25-34"] += 1
            elif 35 <= age <= 44:
                age_group_counts["35-44"] += 1
            else:
                age_group_counts["45+"] += 1

        data = {
            "by_sex": [
                {"label": label, "count": count}
                for label, count in sex_counts.items()
            ],
            "by_education": [
                {"label": label, "count": count}
                for label, count in education_counts.items()
            ],
            "by_age_group": [
                {"age_group": age_group, "count": count}
                for age_group, count in age_group_counts.items()
                if count > 0
            ],
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] applicants_demographics error:", exc)
        return jsonify({"success": False, "message": "Failed to load applicant demographics"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/applicants/location", methods=["GET"])
def applicants_location():
    """Applicants by top cities and is_from_lipa status."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_applicants_filters(request.args, alias="a")

        # Get all applicants with location data and age
        rows = run_query(
            conn,
            f"""
            SELECT
              a.applicant_id,
              COALESCE(NULLIF(a.city, ''), 'Unspecified') AS city,
              a.is_from_lipa,
              a.age
            FROM applicants a
            WHERE {where_sql}
            """,
            params,
            fetch="all",
        ) or []

        # Apply age bracket filtering in Python
        age_brackets = _parse_multi(request.args, "age_bracket")
        if age_brackets:
            rows = [row for row in rows if _matches_age_bracket(
                row['age'], age_brackets)]

        # Calculate city counts from filtered data
        city_counts = {}
        lipa_counts = {"From Lipa": 0, "Not From Lipa": 0}

        for row in rows:
            # City counts
            city = row.get("city")
            city_counts[city] = city_counts.get(city, 0) + 1

            # Lipa status counts
            is_from_lipa = row.get("is_from_lipa")
            if is_from_lipa == 1:
                lipa_counts["From Lipa"] += 1
            else:
                lipa_counts["Not From Lipa"] += 1

        # Get top 10 cities
        top_cities = sorted(city_counts.items(),
                            key=lambda x: x[1], reverse=True)[:10]

        data = {
            "by_city": [
                {"city": city, "count": count}
                for city, count in top_cities
            ],
            "by_is_from_lipa": [
                {"status": status, "count": count}
                for status, count in lipa_counts.items()
            ],
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] applicants_location error:", exc)
        return jsonify({"success": False, "message": "Failed to load applicant location data"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/applicants/experience", methods=["GET"])
def applicants_experience():
    """Applicants by years of experience."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_applicants_filters(request.args, alias="a")

        print(f"DEBUG - WHERE clause: {where_sql}")
        print(f"DEBUG - Params: {params}")

        # Get all applicants first
        rows = run_query(
            conn,
            f"""
            SELECT
              a.applicant_id,
              a.years_experience,
              a.age
            FROM applicants a
            WHERE {where_sql}
            """,
            params,
            fetch="all",
        ) or []

        print(f"DEBUG - Rows before age filtering: {len(rows)}")
        for i, row in enumerate(rows[:5]):  # Show first 5 rows
            print(
                f"  Row {i}: age={row.get('age')}, years_experience='{row.get('years_experience')}'")

        # Apply age bracket filtering in Python
        age_brackets = _parse_multi(request.args, "age_bracket")
        print(f"DEBUG - Age brackets from request: {age_brackets}")

        if age_brackets:
            filtered_rows = []
            for row in rows:
                age = row.get('age')
                matches = _matches_age_bracket(age, age_brackets)
                print(
                    f"  Age {age} matches brackets {age_brackets}: {matches}")
                if matches:
                    filtered_rows.append(row)
            rows = filtered_rows

        print(f"DEBUG - Rows after age filtering: {len(rows)}")

        # Calculate experience ranges from filtered data
        exp_ranges = {
            "No Experience": 0,
            "1-2 Years": 0,
            "3-5 Years": 0,
            "6-10 Years": 0,
            "10+ Years": 0,
            "Unspecified": 0
        }

        for row in rows:
            years_exp = row.get("years_experience")
            print(f"DEBUG - Processing: years_experience='{years_exp}'")

            # Handle your actual data formats
            if (years_exp is None or
                years_exp == '' or
                years_exp == '0' or
                    str(years_exp).lower() == 'none'):
                exp_ranges["No Experience"] += 1

            # Handle range formats like "1-2", "3-4"
            elif isinstance(years_exp, str) and '-' in years_exp:
                try:
                    start, end = years_exp.split('-')
                    start_num = int(start.strip())
                    end_num = int(end.strip())

                    if 1 <= start_num <= 2 and 1 <= end_num <= 2:
                        exp_ranges["1-2 Years"] += 1
                    elif 3 <= start_num <= 5 and 3 <= end_num <= 5:
                        exp_ranges["3-5 Years"] += 1
                    elif 6 <= start_num <= 10 and 6 <= end_num <= 10:
                        exp_ranges["6-10 Years"] += 1
                    else:
                        exp_ranges["Unspecified"] += 1
                except (ValueError, TypeError):
                    exp_ranges["Unspecified"] += 1

            # Handle plus formats like "5+"
            elif isinstance(years_exp, str) and years_exp.endswith('+'):
                try:
                    years_num = int(years_exp.replace('+', '').strip())
                    if 1 <= years_num <= 2:
                        exp_ranges["1-2 Years"] += 1
                    elif 3 <= years_num <= 5:
                        exp_ranges["3-5 Years"] += 1
                    elif 6 <= years_num <= 10:
                        exp_ranges["6-10 Years"] += 1
                    elif years_num > 10:
                        exp_ranges["10+ Years"] += 1
                    else:
                        exp_ranges["Unspecified"] += 1
                except (ValueError, TypeError):
                    exp_ranges["Unspecified"] += 1

            # Handle single number formats
            elif years_exp in ('1', '2'):
                exp_ranges["1-2 Years"] += 1
            elif years_exp in ('3', '4', '5'):
                exp_ranges["3-5 Years"] += 1
            elif years_exp in ('6', '7', '8', '9', '10'):
                exp_ranges["6-10 Years"] += 1

            else:
                try:
                    # Handle any other numeric values
                    years_int = int(years_exp)
                    if years_int == 0:
                        exp_ranges["No Experience"] += 1
                    elif 1 <= years_int <= 2:
                        exp_ranges["1-2 Years"] += 1
                    elif 3 <= years_int <= 5:
                        exp_ranges["3-5 Years"] += 1
                    elif 6 <= years_int <= 10:
                        exp_ranges["6-10 Years"] += 1
                    elif years_int > 10:
                        exp_ranges["10+ Years"] += 1
                    else:
                        exp_ranges["Unspecified"] += 1
                except (ValueError, TypeError):
                    exp_ranges["Unspecified"] += 1

        print(f"DEBUG - Final experience counts: {exp_ranges}")

        data = {
            "by_experience": [
                {"range": exp_range, "count": count}
                for exp_range, count in exp_ranges.items()
                if count > 0
            ],
        }

        # Sort by predefined order
        order = {"No Experience": 1, "1-2 Years": 2,
                 "3-5 Years": 3, "6-10 Years": 4, "10+ Years": 5}
        data["by_experience"].sort(key=lambda x: order.get(x["range"], 6))

        print(f"DEBUG - Final data to return: {data}")

        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] applicants_experience error:", exc)
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": "Failed to load applicant experience data"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/applicants/pwd", methods=["GET"])
def applicants_pwd():
    """Applicants by PWD type."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_applicants_filters(request.args, alias="a")

        # Get all PWD applicants with their type and age
        rows = run_query(
            conn,
            f"""
            SELECT
              a.applicant_id,
              COALESCE(NULLIF(a.pwd_type, ''), 'Not Specified') AS pwd_type,
              a.age
            FROM applicants a
            WHERE {where_sql} AND a.is_pwd = 1
            """,
            params,
            fetch="all",
        ) or []

        # Apply age bracket filtering in Python
        age_brackets = _parse_multi(request.args, "age_bracket")
        if age_brackets:
            rows = [row for row in rows if _matches_age_bracket(
                row['age'], age_brackets)]

        # Calculate PWD type counts from filtered data
        pwd_counts = {}
        for row in rows:
            pwd_type = row.get("pwd_type")
            pwd_counts[pwd_type] = pwd_counts.get(pwd_type, 0) + 1

        data = {
            "by_pwd_type": [
                {"pwd_type": pwd_type, "count": count}
                for pwd_type, count in pwd_counts.items()
            ],
        }
        # Sort by count descending
        data["by_pwd_type"].sort(key=lambda x: x["count"], reverse=True)

        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] applicants_pwd error:", exc)
        return jsonify({"success": False, "message": "Failed to load PWD data"}), 500
    finally:
        conn.close()


# ========== WIDGETS API ENDPOINTS ==========

@admin_bp.route("/api/widgets/notes", methods=["GET"])
def get_notes():
    """Get all notes for the current admin."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        admin_id = session.get("admin_id")
        rows = run_query(
            conn,
            """
            SELECT note_id, title, content, is_pinned, created_at, updated_at
            FROM admin_notes
            WHERE admin_id = %s
            ORDER BY is_pinned DESC, updated_at DESC
            """,
            (admin_id,),
            fetch="all",
        ) or []

        notes = [
            {
                "id": row.get("note_id"),
                "title": row.get("title") or "",
                "content": row.get("content") or "",
                "is_pinned": bool(row.get("is_pinned")),
                "created_at": row.get("created_at").isoformat() if row.get("created_at") else None,
                "updated_at": row.get("updated_at").isoformat() if row.get("updated_at") else None,
            }
            for row in rows
        ]
        return jsonify({"success": True, "data": notes})
    except Exception as exc:
        print("[widgets] get_notes error:", exc)
        return jsonify({"success": False, "message": "Failed to load notes"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/widgets/notes", methods=["POST"])
def create_note():
    """Create a new note."""

    data = request.get_json()
    title = data.get("title", "").strip()
    content = data.get("content", "").strip()
    is_pinned = int(data.get("is_pinned", False))

    if not content:
        return jsonify({"success": False, "message": "Note content is required"}), 400

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        admin_id = session.get("admin_id")
        note_id = run_query(
            conn,
            """
            INSERT INTO admin_notes (admin_id, title, content, is_pinned, created_at, updated_at)
            VALUES (%s, %s, %s, %s, NOW(), NOW())
            """,
            (admin_id, title, content, is_pinned),
            fetch="lastrowid",
        )

        return jsonify({"success": True, "data": {"note_id": note_id}})
    except Exception as exc:
        print("[widgets] create_note error:", exc)
        return jsonify({"success": False, "message": "Failed to create note"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/widgets/notes/<int:note_id>", methods=["PUT"])
def update_note(note_id):
    """Update an existing note."""

    data = request.get_json()
    title = data.get("title", "").strip()
    content = data.get("content", "").strip()
    is_pinned = int(data.get("is_pinned", False))

    if not content:
        return jsonify({"success": False, "message": "Note content is required"}), 400

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        admin_id = session.get("admin_id")
        # Verify ownership
        existing = run_query(
            conn,
            "SELECT note_id FROM admin_notes WHERE note_id = %s AND admin_id = %s",
            (note_id, admin_id),
            fetch="one",
        )
        if not existing:
            return jsonify({"success": False, "message": "Note not found"}), 404

        run_query(
            conn,
            """
            UPDATE admin_notes
            SET title = %s, content = %s, is_pinned = %s, updated_at = NOW()
            WHERE note_id = %s AND admin_id = %s
            """,
            (title, content, is_pinned, note_id, admin_id),
        )

        return jsonify({"success": True})
    except Exception as exc:
        print("[widgets] update_note error:", exc)
        return jsonify({"success": False, "message": "Failed to update note"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/widgets/notes/<int:note_id>", methods=["DELETE"])
def delete_note(note_id):
    """Delete a note."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        admin_id = session.get("admin_id")
        # Verify ownership
        existing = run_query(
            conn,
            "SELECT note_id FROM admin_notes WHERE note_id = %s AND admin_id = %s",
            (note_id, admin_id),
            fetch="one",
        )
        if not existing:
            return jsonify({"success": False, "message": "Note not found"}), 404

        run_query(
            conn,
            "DELETE FROM admin_notes WHERE note_id = %s AND admin_id = %s",
            (note_id, admin_id),
        )

        return jsonify({"success": True})
    except Exception as exc:
        print("[widgets] delete_note error:", exc)
        return jsonify({"success": False, "message": "Failed to delete note"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/widgets/preferences", methods=["GET"])
def get_widget_preferences():
    """Get widget visibility preferences for the current admin."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        admin_id = session.get("admin_id")
        row = run_query(
            conn,
            "SELECT widget_preferences FROM admin_widget_preferences WHERE admin_id = %s",
            (admin_id,),
            fetch="one",
        )

        if row and row.get("widget_preferences"):
            prefs = json.loads(row.get("widget_preferences"))
        else:
            # Default: all widgets visible
            prefs = {
                "clock": True,
                "calendar": True,
                "notes": True,
                "todo": True,
                "countdown": True,
                "quickLinks": True,
                "quotes": True,
                "notifications": True,
                "productivity": True,
            }

        return jsonify({"success": True, "data": prefs})
    except Exception as exc:
        print("[widgets] get_widget_preferences error:", exc)
        return jsonify({"success": False, "message": "Failed to load preferences"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/widgets/preferences/reset", methods=["POST"])
def reset_widget_preferences():
    """Reset widget preferences to default for the current admin."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        admin_id = session.get("admin_id")

        # Default: all widgets visible
        default_prefs = {
            "clock": True,
            "calendar": True,
            "notes": True,
            "todo": True,
            "countdown": True,
            "quickLinks": True,
            "quotes": True,
            "notifications": True,
            "productivity": True,
        }

        default_json = json.dumps(default_prefs)

        # Save into the DB (overwrite existing)
        run_query(
            conn,
            """
            INSERT INTO admin_widget_preferences (admin_id, widget_preferences, updated_at)
            VALUES (%s, %s, NOW())
            ON DUPLICATE KEY UPDATE widget_preferences = %s, updated_at = NOW()
            """,
            (admin_id, default_json, default_json),
        )

        # Return for frontend
        return jsonify({"success": True, "data": default_prefs})
    except Exception as exc:
        print("[widgets] reset_widget_preferences error:", exc)
        return jsonify({"success": False, "message": "Failed to reset preferences"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/widgets/preferences", methods=["POST"])
def save_widget_preferences():
    """Save widget visibility preferences."""

    data = request.get_json()
    preferences = data.get("preferences", {})

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        admin_id = session.get("admin_id")
        prefs_json = json.dumps(preferences)

        # Insert or update
        run_query(
            conn,
            """
            INSERT INTO admin_widget_preferences (admin_id, widget_preferences, updated_at)
            VALUES (%s, %s, NOW())
            ON DUPLICATE KEY UPDATE widget_preferences = %s, updated_at = NOW()
            """,
            (admin_id, prefs_json, prefs_json),
        )

        return jsonify({"success": True})
    except Exception as exc:
        print("[widgets] save_widget_preferences error:", exc)
        return jsonify({"success": False, "message": "Failed to save preferences"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/widgets/weather", methods=["GET"])
def get_weather():
    """Get weather data using Open-Meteo API."""
    city = request.args.get("city", "Lipa City")

    # Coordinates for Lipa, Batangas
    latitude = 13.9444
    longitude = 121.1631

    try:
        url = "https://api.open-meteo.com/v1/forecast"

        params = {
            "latitude": latitude,
            "longitude": longitude,
            "current_weather": True,
            "timezone": "Asia/Manila",
        }

        headers = {
            "User-Agent": "DashboardWidget/1.0"  # REQUIRED by Open-Meteo
        }

        response = requests.get(url, params=params, headers=headers, timeout=7)
        response.raise_for_status()

        data = response.json()

        if "current_weather" not in data:
            raise Exception("No current weather field in API response")

        current = data["current_weather"]

        # Extract values safely
        temperature = current.get("temperature")
        wind_speed = current.get("windspeed")
        weather_code = current.get("weathercode")

        # Your weather code mapper function
        weather_info = get_weather_info(weather_code)

        return jsonify({
            "success": True,
            "data": {
                "city": city,
                "temperature": temperature,
                "description": weather_info["description"],
                "icon": weather_info["icon"],
                "wind_speed": wind_speed,
                "weather_code": weather_code
            }
        })

    except Exception as e:
        print(f"[widgets] Weather error: {e}")

        # Fallback
        return jsonify({
            "success": True,
            "data": {
                "city": city,
                "temperature": 28,
                "description": "Weather unavailable",
                "icon": "cloud",
                "wind_speed": 0
            }
        })


def get_weather_info(weather_code):
    """Map Open-Meteo weather codes to descriptions and icons.

    Weather codes from Open-Meteo:
    0 = Clear sky
    1-3 = Mainly clear, partly cloudy, overcast
    45-48 = Fog
    51-67 = Drizzle and rain
    71-77 = Snow
    80-82 = Rain showers
    85-86 = Snow showers
    95-99 = Thunderstorm
    """
    code = int(weather_code) if weather_code else 0

    if code == 0:
        return {"description": "Clear Sky", "icon": "sun"}
    elif code in [1, 2, 3]:
        return {"description": "Partly Cloudy", "icon": "cloud-sun"}
    elif code in [45, 48]:
        return {"description": "Foggy", "icon": "smog"}
    elif code in [51, 53, 55]:
        return {"description": "Light Drizzle", "icon": "cloud-rain"}
    elif code in [56, 57]:
        return {"description": "Freezing Drizzle", "icon": "snowflake"}
    elif code in [61, 63, 65]:
        return {"description": "Rain", "icon": "cloud-rain"}
    elif code in [66, 67]:
        return {"description": "Freezing Rain", "icon": "snowflake"}
    elif code in [71, 73, 75, 77]:
        return {"description": "Snow", "icon": "snowflake"}
    elif code in [80, 81, 82]:
        return {"description": "Rain Showers", "icon": "cloud-showers-heavy"}
    elif code in [85, 86]:
        return {"description": "Snow Showers", "icon": "snowflake"}
    elif code in [95, 96, 99]:
        return {"description": "Thunderstorm", "icon": "bolt"}
    else:
        return {"description": "Unknown", "icon": "cloud"}


@admin_bp.route("/api/widgets/productivity", methods=["GET"])
def get_productivity_stats():
    """Get productivity statistics for today."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        today = datetime.today().date()

        # Applications reviewed today (use applied_at since updated_at doesn't exist)
        apps_reviewed = run_query(
            conn,
            """
            SELECT COUNT(*) AS count
            FROM applications
            WHERE DATE(applied_at) = %s AND status IN ('Approved', 'Rejected')
            """,
            (today,),
            fetch="one",
        ) or {}

        # New registrations today
        new_regs = run_query(
            conn,
            """
            SELECT COUNT(*) AS count
            FROM applicants
            WHERE DATE(created_at) = %s
            """,
            (today,),
            fetch="one",
        ) or {}

        # New employers registered today
        new_employers = run_query(
            conn,
            """
            SELECT COUNT(*) AS count
            FROM employers
            WHERE DATE(created_at) = %s
            """,
            (today,),
            fetch="one",
        ) or {}

        data = {
            # Use applications reviewed as tasks
            "tasks_completed": _to_int(apps_reviewed.get("count")),
            "applications_reviewed": _to_int(apps_reviewed.get("count")),
            "new_registrations": _to_int(new_regs.get("count")),
            "new_employers": _to_int(new_employers.get("count")),
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[widgets] get_productivity_stats error:", exc)
        return jsonify({"success": False, "message": "Failed to load productivity stats"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/employers/summary", methods=["GET"])
def employers_summary():
    """Employers volume & active vs inactive, respecting filters."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_employers_filters(request.args, alias="e")
        row = run_query(
            conn,
            f"""
            SELECT
              COUNT(*) AS total_employers,
              SUM(CASE WHEN e.is_active = 1 THEN 1 ELSE 0 END) AS active_count,
              SUM(CASE WHEN e.is_active = 0 THEN 1 ELSE 0 END) AS inactive_count
            FROM employers e
            WHERE {where_sql}
            """,
            params,
            fetch="one",
        ) or {}

        total_employers = _to_int(row.get("total_employers"))
        active_count = _to_int(row.get("active_count"))
        inactive_count = _to_int(row.get("inactive_count"))

        # Pending documents: approximate as status = 'Pending'
        pending_row = run_query(
            conn,
            f"""
            SELECT COUNT(*) AS pending_docs
            FROM employers e
            WHERE {where_sql} AND e.status = 'Pending'
            """,
            params,
            fetch="one",
        ) or {}

        data = {
            "total_employers": total_employers,
            "active_employers": active_count,
            "inactive_employers": inactive_count,
            "pending_documents": _to_int(pending_row.get("pending_docs")),
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] employers_summary error:", exc)
        return jsonify({"success": False, "message": "Failed to load employers summary"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/employers/business", methods=["GET"])
def employers_business():
    """Employers by industry and recruitment type, respecting filters."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_employers_filters(request.args, alias="e")

        industry_rows = run_query(
            conn,
            f"""
            SELECT
              COALESCE(NULLIF(e.industry, ''), 'Unspecified') AS industry,
              COUNT(*) AS total
            FROM employers e
            WHERE {where_sql}
            GROUP BY industry
            ORDER BY total DESC
            """,
            params,
            fetch="all",
        ) or []

        rec_rows = run_query(
            conn,
            f"""
            SELECT
              COALESCE(NULLIF(e.recruitment_type, ''), 'Unspecified') AS recruitment_type,
              COUNT(*) AS total
            FROM employers e
            WHERE {where_sql}
            GROUP BY recruitment_type
            """,
            params,
            fetch="all",
        ) or []

        data = {
            "by_industry": [
                {"industry": r.get("industry"),
                 "count": _to_int(r.get("total"))}
                for r in industry_rows
            ],
            "by_recruitment_type": [
                {
                    "recruitment_type": r.get("recruitment_type"),
                    "count": _to_int(r.get("total")),
                }
                for r in rec_rows
            ],
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] employers_business error:", exc)
        return jsonify({"success": False, "message": "Failed to load employer demographics"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/employers/location", methods=["GET"])
def employers_location():
    """Employers by top cities and provinces."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_employers_filters(request.args, alias="e")

        # Top 10 cities
        city_rows = run_query(
            conn,
            f"""
            SELECT
              COALESCE(NULLIF(e.city, ''), 'Unspecified') AS city,
              COUNT(*) AS total
            FROM employers e
            WHERE {where_sql}
            GROUP BY city
            ORDER BY total DESC
            LIMIT 10
            """,
            params,
            fetch="all",
        ) or []

        # Top 10 provinces
        province_rows = run_query(
            conn,
            f"""
            SELECT
              COALESCE(NULLIF(e.province, ''), 'Unspecified') AS province,
              COUNT(*) AS total
            FROM employers e
            WHERE {where_sql}
            GROUP BY province
            ORDER BY total DESC
            LIMIT 10
            """,
            params,
            fetch="all",
        ) or []

        data = {
            "by_city": [
                {"city": r.get("city"), "count": _to_int(r.get("total"))}
                for r in city_rows
            ],
            "by_province": [
                {"province": r.get("province"),
                 "count": _to_int(r.get("total"))}
                for r in province_rows
            ],
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] employers_location error:", exc)
        return jsonify({"success": False, "message": "Failed to load employer location data"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/employers/status", methods=["GET"])
def employers_status():
    """Employers by status."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_employers_filters(request.args, alias="e")

        status_rows = run_query(
            conn,
            f"""
            SELECT
              COALESCE(NULLIF(e.status, ''), 'Unspecified') AS status,
              COUNT(*) AS total
            FROM employers e
            WHERE {where_sql}
            GROUP BY status
            ORDER BY total DESC
            """,
            params,
            fetch="all",
        ) or []

        data = {
            "by_status": [
                {"status": r.get("status"), "count": _to_int(r.get("total"))}
                for r in status_rows
            ],
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] employers_status error:", exc)
        return jsonify({"success": False, "message": "Failed to load employer status data"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/jobs/summary", methods=["GET"])
def jobs_summary():
    """Job demand KPIs: total open jobs."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_jobs_filters(request.args, alias="j")

        # For "open jobs" metric, we want active jobs that haven't expired
        # Check if user has specifically filtered for job status
        job_statuses = _parse_multi(request.args, "job_status")

        if not job_statuses:
            # No job status filter applied - show only active, non-expired jobs for "open jobs" metric
            open_jobs_where = f"({where_sql}) AND j.status = 'active' AND (j.job_expiration_date IS NULL OR j.job_expiration_date >= CURDATE())"
        else:
            # User applied job status filters - respect their selection but still check expiration for "open jobs"
            open_jobs_where = f"({where_sql}) AND (j.job_expiration_date IS NULL OR j.job_expiration_date >= CURDATE())"

        row = run_query(
            conn,
            f"""
            SELECT COUNT(*) AS open_jobs
            FROM jobs j
            WHERE {open_jobs_where}
            """,
            params,
            fetch="one",
        ) or {}

        data = {
            "total_open_jobs": _to_int(row.get("open_jobs")),
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] jobs_summary error:", exc)
        return jsonify({"success": False, "message": "Failed to load jobs summary"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/jobs/demand", methods=["GET"])
def jobs_demand():
    """Top job positions and jobs by work schedule."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_jobs_filters(request.args, alias="j")

        pos_rows = run_query(
            conn,
            f"""
            SELECT
              COALESCE(NULLIF(j.job_position, ''), 'Unspecified') AS job_position,
              COUNT(*) AS total
            FROM jobs j
            WHERE {where_sql}
            GROUP BY job_position
            ORDER BY total DESC
            LIMIT 5
            """,
            params,
            fetch="all",
        ) or []

        sched_rows = run_query(
            conn,
            f"""
            SELECT
              COALESCE(NULLIF(j.work_schedule, ''), 'Unspecified') AS work_schedule,
              COUNT(*) AS total
            FROM jobs j
            WHERE {where_sql}
            GROUP BY work_schedule
            """,
            params,
            fetch="all",
        ) or []

        data = {
            "by_position": [
                {"job_position": r.get("job_position"),
                 "count": _to_int(r.get("total"))}
                for r in pos_rows
            ],
            "by_work_schedule": [
                {"work_schedule": r.get("work_schedule"),
                 "count": _to_int(r.get("total"))}
                for r in sched_rows
            ],
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] jobs_demand error:", exc)
        return jsonify({"success": False, "message": "Failed to load job demand"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/applications/summary", methods=["GET"])
def applications_summary():
    """Applications flow KPIs: total applications, status breakdown, success rate."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_applications_filters(request.args, alias="a")

        rows = run_query(
            conn,
            f"""
            SELECT
              COALESCE(NULLIF(a.status, ''), 'Pending') AS status,
              COUNT(*) AS total
            FROM applications a
            LEFT JOIN jobs j ON a.job_id = j.job_id
            WHERE {where_sql}
            GROUP BY status
            """,
            params,
            fetch="all",
        ) or []

        by_status = [
            {"status": r.get("status"), "count": _to_int(r.get("total"))}
            for r in rows
        ]

        total_applications = sum(item["count"] for item in by_status)
        # Update success rate calculation to use "Hired" instead of "Approved"
        hired = next(
            (item["count"]
             for item in by_status if item["status"] == "Hired"), 0
        )
        success_rate = (
            hired / total_applications) if total_applications else 0

        data = {
            "total_applications": total_applications,
            "by_status": by_status,
            "success_rate": success_rate,
            "success_rate_percentage": round(success_rate * 100, 2)
            if total_applications
            else 0,
        }
        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] applications_summary error:", exc)
        return jsonify({"success": False, "message": "Failed to load applications summary"}), 500
    finally:
        conn.close()


@admin_bp.route("/api/analytics/applications/trend", methods=["GET"])
def applications_trend():
    """Applications by month/year trend."""

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    try:
        where_sql, params = build_applications_filters(request.args, alias="a")

        rows = run_query(
            conn,
            f"""
            SELECT
            DATE_FORMAT(a.applied_at, '%b %Y') AS label,
            YEAR(a.applied_at) AS year_num,
            MONTH(a.applied_at) AS month_num,
            COUNT(*) AS total
            FROM applications a
            LEFT JOIN jobs j ON a.job_id = j.job_id
            WHERE {where_sql}
            GROUP BY year_num, month_num
            ORDER BY year_num ASC, month_num ASC
            """,
            params,
            fetch="all",
        ) or []

        data = [
            {
                "label": r.get("label"),
                "year": _to_int(r.get("year_num")),
                "month": _to_int(r.get("month_num")),
                "count": _to_int(r.get("total")),
            }
            for r in rows
        ]

        # Check if we have meaningful data (same logic as other charts)
        has_data = len(data) > 0 and any(item["count"] > 0 for item in data)

        if not has_data:
            # Return empty array to trigger the "no data" state in frontend
            return jsonify({"success": True, "data": []})

        return jsonify({"success": True, "data": data})
    except Exception as exc:
        print("[analytics] applications_trend error:", exc)
        return jsonify({"success": False, "message": "Failed to load applications trend"}), 500
    finally:
        conn.close()


# ===== API: Get Notifications =====
@admin_bp.route("/api/notifications", methods=["GET"])
def api_get_notifications():
    filter_param = request.args.get("filter", "all")
    notification_type = request.args.get("type")

    is_read = None
    if filter_param == "read":
        is_read = True
    elif filter_param == "unread":
        is_read = False

    # 1. Fetch ALL notifications first (we will filter in Python)
    # We don't use exclude_types here anymore to avoid leaking new types
    all_notifications = get_notifications(
        notification_type=notification_type,
        is_read=is_read
    )

    # 2. Define STRICT list of Admin-only notification types
    admin_types = {
        'applicant_approval',
        'employer_approval',
        'applicant_reported',
        'employer_reported',
        'applicant_outdated_docu',
        'employer_outdated_docu',
        'applicant_batch'
    }

    # 3. Filter: Keep only notifications that belong to Admin
    final_notifications = [
        n for n in all_notifications
        if n.get('notification_type') in admin_types
    ]

    return jsonify({
        "success": True,
        "notifications": final_notifications,
        "count": len(final_notifications)
    })

# ===== API: Mark Notification as Read =====


@admin_bp.route("/api/notifications/<int:notification_id>/read", methods=["POST"])
def api_mark_notification_read(notification_id):

    result = mark_notification_read(notification_id)

    if result:
        return jsonify({"success": True, "message": "Notification marked as read"})
    else:
        return jsonify({"success": False, "message": "Failed to mark notification as read"}), 500


# ===== API: Get Unread Count =====
@admin_bp.route("/api/notifications/unread-count", methods=["GET"])
def api_unread_count():
    # Use the same strict Allowlist to ensure the dot matches the list
    admin_types = [
        'applicant_approval',
        'employer_approval',
        'applicant_reported',
        'employer_reported',
        'applicant_outdated_docu',
        'employer_outdated_docu',
        'applicant_batch'
    ]

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "unread_count": 0})

    try:
        # Dynamically build the placeholder string based on the list length
        placeholders = ', '.join(['%s'] * len(admin_types))

        query = f"""
            SELECT COUNT(*) as count 
            FROM notifications 
            WHERE is_read = 0 
            AND notification_type IN ({placeholders})
        """

        # Pass the list as parameters
        result = run_query(conn, query, tuple(admin_types), fetch="one")
        count = result['count'] if result else 0

        return jsonify({"success": True, "unread_count": count})
    except Exception as e:
        print(f"[admin] Error counting unread: {e}")
        return jsonify({"success": False, "unread_count": 0})
    finally:
        conn.close()


@admin_bp.route("/approve-reupload/<int:applicant_id>", methods=["POST"])
def approve_reupload(applicant_id):
    try:
        conn = create_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM applicants WHERE applicant_id = %s", (applicant_id,))
        applicant = cursor.fetchone()

        if not applicant:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Applicant not found"}), 404

        # Update status to Approved
        cursor.execute(
            "UPDATE applicants SET status = %s WHERE applicant_id = %s",
            ("Approved", applicant_id)
        )
        conn.commit()

        # Send approval email
        subject = "PESO SmartHire - Full Access Granted"
        body = f"""
        <p>Hi {applicant['first_name']},</p>
        <p>This is PESO SmartHire Team.</p>
        <p>Congratulations! Your reuploaded document has been reviewed and approved.</p>
        <p>You now have full access to all features of the PESO SmartHire platform.</p>
        <p>Thank you for your cooperation!</p>
        <p>— PESO SmartHire Admin</p>
        """
        msg = Message(subject=subject, recipients=[
                      applicant["email"]], html=body)
        mail.send(msg)

        cursor.close()
        conn.close()

        # Also flash server-side so the message is visible after reload
        flash("Applicant approved and notified successfully", "success")
        return jsonify({"success": True, "message": "Applicant approved and notified successfully"})

    except Exception as e:
        print(f"[v0] Error approving reupload: {e}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        flash(str(e), "danger")
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route("/approve-employer-reupload/<int:employer_id>", methods=["POST"])
def approve_employer_reupload(employer_id):
    try:
        conn = create_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM employers WHERE employer_id = %s", (employer_id,))
        employer = cursor.fetchone()

        if not employer:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Employer not found"}), 404

        # Update status to Approved
        cursor.execute(
            "UPDATE employers SET status = %s WHERE employer_id = %s",
            ("Approved", employer_id)
        )
        conn.commit()

        # Send approval email
        subject = "PESO SmartHire - Full Access Granted"
        body = f"""
        <p>Dear {employer['employer_name']},</p>
        <p>This is PESO SmartHire Team.</p>
        <p>Congratulations! Your reuploaded documents have been reviewed and approved.</p>
        <p>You now have full access to all features of the PESO SmartHire employer platform.</p>
        <p>Thank you for your cooperation!</p>
        <p>— PESO SmartHire Admin</p>
        """
        msg = Message(subject=subject, recipients=[
                      employer["email"]], html=body)
        mail.send(msg)

        cursor.close()
        conn.close()

        flash("Employer approved and notified successfully", "success")
        return jsonify({"success": True, "message": "Employer approved and notified successfully"})

    except Exception as e:
        print(f"[v0] Error approving employer reupload: {e}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        flash(str(e), "danger")
        return jsonify({"success": False, "message": str(e)}), 500


# ==========================
# UPDATE NON-LIPEÑO STATUS + SEND EMAIL
# ==========================
@admin_bp.route("/update_nonlipeno_status/<int:applicant_id>", methods=["POST"])
def update_nonlipeno_status(applicant_id):
    try:
        data = request.get_json()
        print(f"[v1] Received data for applicant {applicant_id}: {data}")

        if not data or "action" not in data:
            print("[v1] No action provided in request")
            return jsonify({"success": False, "message": "No action provided."}), 400

        action = data["action"]
        reason = None

        conn = create_connection()
        if not conn:
            print("[v1] Database connection failed")
            return jsonify({"success": False, "message": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM applicants WHERE applicant_id = %s AND is_from_lipa = 0",
            (applicant_id,)
        )
        applicant = cursor.fetchone()

        if not applicant:
            print(f"[v1] Non-Lipeño applicant {applicant_id} not found in DB")
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Non-Lipeño applicant not found"}), 404

        # If must_change_password = 1, this is a NEW applicant who hasn't changed their password yet
        # If must_change_password = 0, this is an EXISTING applicant (already changed password once)
        is_new_applicant = applicant.get("must_change_password") == 1

        print(f"[v0] User detection for applicant {applicant_id}:")
        print(
            f"- must_change_password: {applicant.get('must_change_password')}")
        print(f"- Is new applicant: {is_new_applicant}")
        print(f"- Current status: {applicant.get('status')}")

        temp_password_plain = None

        if action == "approved":
            if is_new_applicant:
                # Only generate new credentials for first-time approval
                print(
                    f"[v0] NEW applicant {applicant_id} - generating credentials with must_change_password=1")
                temp_password_plain = secrets.token_urlsafe(8)
                password_hash = generate_password_hash(temp_password_plain)
                cursor.execute(
                    "UPDATE applicants SET password_hash = %s, temp_password = %s, must_change_password = 1 WHERE applicant_id = %s",
                    (password_hash, temp_password_plain, applicant_id)
                )
                applicant["temp_password"] = temp_password_plain
            else:
                # EXISTING applicant - keep current credentials and DO NOT reset must_change_password
                print(
                    f"[v0] EXISTING applicant {applicant_id} - keeping current credentials, NOT resetting must_change_password")
                temp_password_plain = applicant.get("temp_password")

        elif action == "reupload":
            # For reupload - use existing credentials
            print(
                f"[v0] Reupload request - using existing credentials for applicant {applicant_id}")
            temp_password_plain = applicant.get("temp_password")

        # EXISTING users keep their must_change_password = 0 (they already changed password)

        if action == "approved":
            new_status = "Approved"
            subject = "PESO SmartHire - Application Approved"

            if is_new_applicant:
                # NEW applicant - send email with credentials
                body = f"""
                <p>Hi {applicant['first_name']},</p>
                <p>This is PESO SmartHire Team.</p>
                <p>Congratulations! Your registration has been reviewed and approved!</p>
                <p>Included below are your login credentials:</p>
                <ul>
                    <li>Applicant ID: {applicant['applicant_code']}</li>
                    <li>Email: {applicant['email']}</li>
                    <li>Phone Number: {applicant['phone']}</li>
                    <li>Password: {temp_password_plain}</li>
                </ul>
                <p><strong>Please change your password after logging in.</strong></p>
                <p>Thank you for joining our PESO SmartHire Platform.</p>
                <p>— PESO SmartHire Admin</p>
                """
            else:
                # EXISTING applicant (residency change) - send approval without credentials
                body = f"""
                <p>Hi {applicant['first_name']},</p>
                <p>This is PESO SmartHire Team.</p>
                <p>Congratulations! Your residency change has been reviewed and approved.</p>
                <p>You now have full access to all features of the PESO SmartHire platform.</p>
                <p>You can log in using your existing credentials to continue using our services.</p>
                <p>Thank you for keeping your information up to date!</p>
                <p>— PESO SmartHire Admin</p>
                """

            success_message = "Non-Lipeño applicant approved successfully! Credentials sent via email."

        elif action == "rejected":
            reason = data.get("reason")
            subject = "PESO SmartHire - Application Status Update"
            reason_block = f"<p><strong>Reason:</strong> {reason}</p>" if reason else ""

            body = f"""
            <p>Hi {applicant['first_name']},</p>
            <p>This is PESO SmartHire Team.</p>
            <p>We regret to inform you that your application for PESO SmartHire has been reviewed but did not meet the current requirements.</p>
            {reason_block}
            <p>You may reapply in the future once you meet the qualifications.</p>
            <p>— PESO SmartHire Admin</p>
            """

            # 1. Send Email FIRST (before deleting data)
            msg = Message(subject=subject, recipients=[
                          applicant["email"]], html=body)
            try:
                mail.send(msg)
            except Exception as e:
                logger.error(f"Failed to send rejection email: {e}")

            # 2. Delete Associated Files (Clean up storage)
            file_fields = ["recommendation_letter_path",
                           "profile_picture_path", "resume_path"]
            for field in file_fields:
                file_path = applicant.get(field)
                if file_path:
                    try:
                        full_path = os.path.join("static", file_path)
                        if os.path.exists(full_path):
                            os.remove(full_path)
                    except Exception as e:
                        print(f"Error deleting file {file_path}: {e}")

            # 3. DELETE the record from Database
            cursor.execute(
                "DELETE FROM applicants WHERE applicant_id = %s", (applicant_id,))
            conn.commit()

            cursor.close()
            conn.close()
            return jsonify({"success": True, "message": "Applicant rejected and account deleted."})

        elif action == "reupload":
            new_status = "Reupload"
            document_name = data.get("document_name", "Recommendation Letter")
            subject = "PESO SmartHire - Document Reupload Required"

            if is_new_applicant:
                # NEW applicant - send welcome email with credentials and reupload instructions
                body = f"""
                <p>Hi {applicant['first_name']},</p>
                <p>This is PESO SmartHire Team.</p>
                <p>We have reviewed your application for PESO SmartHire. To proceed with your application, we need you to upload your {document_name}.</p>
                <p>To help you get started, here are your login credentials:</p>
                <ul>
                    <li>Applicant ID: {applicant['applicant_code']}</li>
                    <li>Email: {applicant['email']}</li>
                    <li>Phone Number: {applicant['phone']}</li>
                    <li>Password: {temp_password_plain}</li>
                </ul>
                <p><strong>Steps to Upload Your Document:</strong></p>
                <ol>
                    <li>Log in to your account using the credentials above</li>
                    <li>Upload your {document_name}</li>
                </ol>
                <p>We'll review your document once it's uploaded and notify you of any updates.</p>
                <p>Thank you for choosing PESO SmartHire!</p>
                <p>— PESO SmartHire Admin</p>
                """
            else:
                # EXISTING applicant - regular reupload (don't force password change)
                body = f"""
                <p>Hi {applicant['first_name']},</p>
                <p>This is PESO SmartHire Team.</p>
                <p>We need you to provide an updated {document_name} for your application.</p>
                <p><strong>Required Action:</strong></p>
                <ol>
                    <li>Log in to your PESO SmartHire account</li>
                    <li>Upload your updated {document_name}</li>
                </ol>
                <p>We'll review your document once it's uploaded and update your application status accordingly.</p>
                <p>Note: If you've forgotten your password, you can reset it using the "Forgot Password" option on the login page.</p>
                <p>Thank you for your cooperation!</p>
                <p>— PESO SmartHire Admin</p>
                """

            success_message = "Re-upload request sent. Email notification sent to applicant."

        else:
            print(f"[v1] Invalid action: {action}")
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Invalid action."}), 400

        is_active_value = 1 if new_status == "Approved" else 0

        if reason:
            cursor.execute(
                "UPDATE applicants SET status = %s, rejection_reason = %s, is_active = %s WHERE applicant_id = %s",
                (new_status, reason, is_active_value, applicant_id)
            )
        else:
            cursor.execute(
                "UPDATE applicants SET status = %s, is_active = %s, approved_at = NOW() WHERE applicant_id = %s",
                (new_status, is_active_value, applicant_id)
            )

        conn.commit()

        msg = Message(
            subject=subject,
            recipients=[applicant["email"]],
            html=body
        )
        mail.send(msg)

        cursor.close()
        conn.close()

        print(
            f"[v1] Status updated and email sent for applicant {applicant_id}")
        return jsonify({"success": True, "message": success_message})

    except Exception as e:
        print(f"[v1] Error updating status or sending email: {e}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "message": f"An error occurred: {str(e)}"}), 500


@admin_bp.route("/login", methods=["GET", "POST"])  # 1. Allow GET requests
def login():
    # 2. Handle GET request (Show the HTML Page)
    if request.method == "GET":
        # Points to templates/admin/admin_login.html
        return render_template("admin/admin_login.html")

    # 3. Handle POST request (Process the Login Form)
    admin_code = request.form.get("adminID")
    email = request.form.get("adminEmail")
    password = request.form.get("adminPassword")

    if not admin_code or not email or not password:
        flash("Please fill in all fields.", "danger")
        # Redirect back to login, not home
        return redirect(url_for("admin.login"))

    conn = create_connection()
    if not conn:
        flash("Database connection failed.", "danger")
        return redirect(url_for("admin.login"))

    query = "SELECT * FROM admin WHERE admin_code = %s AND email = %s"
    result = run_query(conn, query, (admin_code, email), fetch="one")
    conn.close()

    if result:
        if check_password_hash(result["password_hash"], password):
            session["admin_id"] = result["admin_id"]
            session["admin_code"] = result["admin_code"]
            session["admin_email"] = result["email"]

            flash("Welcome Back Administrator!", "success")
            return redirect(url_for("admin.admin_home"))
        else:
            flash("Invalid password.", "danger")
    else:
        flash("Invalid Admin ID or Email.", "danger")

    # If login fails, stay on the login page
    return redirect(url_for("admin.login"))


@admin_bp.route("/notifications")
def notifications_page():
    admin_id = session.get("admin_id")
    notifications = get_notifications(
        exclude_types=['job_application', 'report_verdict'])

    return render_template("Admin/admin_notif.html", notifications=notifications)


# ===== Admin: Applicants Management =====
@admin_bp.route("/applicants")
def applicants_management():
    return render_template("Admin/admin_applicant.html")


@admin_bp.route("/applicants/for-approval")
def applicants_for_approval():
    """Show non-Lipeño applicants needing approval"""

    conn = create_connection()
    cursor = conn.cursor(dictionary=True)

    # Get non-Lipeño applicants with Pending status
    cursor.execute("""
        SELECT applicant_id, first_name, middle_name, last_name,
               created_at, status, is_from_lipa
        FROM applicants
        WHERE is_from_lipa = 0 AND status = 'Pending'
        ORDER BY created_at DESC
    """)
    applicants = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("Admin/applicants_for_approval.html", applicants=applicants)


@admin_bp.route("/applicants/view-all")
def applicants_view_all():
    """Show all applicants with Lipeño/Non-Lipeño filter"""

    conn = create_connection()
    cursor = conn.cursor(dictionary=True)

    # Get all applicants (all statuses)
    cursor.execute("""
        SELECT applicant_id, first_name, middle_name, last_name,
               created_at, status, is_from_lipa
        FROM applicants
        ORDER BY created_at DESC
    """)
    applicants = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("Admin/applicants_view_all.html", applicants=applicants)


@admin_bp.route('/applicants_for_reported_acc')
def applicants_for_reported_acc():
    """Show job posts reported by applicants for moderation."""

    conn = create_connection()
    if not conn:
        flash("Database connection failed", "danger")
        return redirect(url_for("admin.admin_home"))

    job_reports = []

    try:
        cursor = conn.cursor(dictionary=True)

        # Fetch job reports only (correct schema using created_at)
        cursor.execute("""
            SELECT 
                jr.id AS report_id,
                jr.applicant_id,
                jr.job_id,
                jr.reason,
                jr.details,
                jr.created_at AS reported_at,

                CONCAT(
                    COALESCE(reporter.first_name, 'Unknown'), ' ',
                    COALESCE(reporter.middle_name, ''), ' ',
                    COALESCE(reporter.last_name, 'Applicant')
                ) AS reported_by_name,

                COALESCE(j.job_position, 'Deleted Job') AS job_title,
                COALESCE(e.employer_name, 'Unknown Employer') AS employer_name,

                e.employer_id,
                COALESCE(jr.status, 'Pending') AS status,
                j.status AS job_status
            FROM job_reports jr
            LEFT JOIN applicants reporter ON jr.applicant_id = reporter.applicant_id
            LEFT JOIN jobs j ON jr.job_id = j.job_id
            LEFT JOIN employers e ON j.employer_id = e.employer_id
            ORDER BY jr.created_at DESC;
        """)

        job_reports = cursor.fetchall() or []

        cursor.close()

    except Exception as e:
        print(f"[ERROR fetching reported job posts: {str(e)}]")
        flash(f"Error loading reported job posts: {str(e)}", "danger")

    finally:
        conn.close()

    return render_template(
        'Admin/applicants_for_reported_acc.html',
        job_reports=job_reports
    )


@admin_bp.route("/reported_applicants")
def reported_applicants():

    conn = create_connection()
    if not conn:
        flash("Database connection failed", "danger")
        return redirect(url_for("admin.employers_management"))

    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                ar.id AS report_id,
                ar.applicant_id,
                ar.employer_id,
                ar.job_id,
                ar.reason,
                ar.details,
                ar.created_at AS reported_at,

                CONCAT(
                    COALESCE(app.first_name, 'Unknown'), ' ',
                    COALESCE(app.last_name, 'Applicant')
                ) AS applicant_name,

                COALESCE(emp.employer_name, 'Unknown Employer') AS employer_name,
                emp.email AS employer_email,
                COALESCE(job.job_position, 'N/A') AS job_title,
                COALESCE(ar.status, 'Pending') AS status
            FROM applicant_reports ar
            LEFT JOIN applicants app ON ar.applicant_id = app.applicant_id
            LEFT JOIN employers emp ON ar.employer_id = emp.employer_id
            LEFT JOIN jobs job ON ar.job_id = job.job_id
            ORDER BY ar.created_at DESC;
        """)
        reports = cursor.fetchall() or []
        cursor.close()
    except Exception as exc:
        print(f"[v1] Failed to load applicant reports: {exc}")
        reports = []
        flash("Unable to load reported applicants.", "danger")
    finally:
        conn.close()

    return render_template(
        "Admin/reported_applicants.html",
        reports=reports
    )


@admin_bp.route("/get_job_details/<int:job_id>")
def get_job_details(job_id):
    try:
        print(f"[DEBUG] Fetching job details for job_id: {job_id}")

        conn = create_connection()
        if not conn:
            print("[DEBUG] Database connection failed")
            return jsonify({"success": False, "message": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        print(f"[DEBUG] Executing SQL query for job_id: {job_id}")

        # CORRECTED QUERY - NO requirements column!
        cursor.execute("""
            SELECT 
                j.job_id,
                j.job_position,
                j.job_description,
                j.qualifications,
                j.work_schedule,
                j.min_salary,
                j.max_salary,
                j.status,
                j.created_at,
                j.job_expiration_date,
                j.application_count,
                e.employer_name,
                e.city,
                e.province,
                e.employer_id
            FROM jobs j
            LEFT JOIN employers e ON j.employer_id = e.employer_id
            WHERE j.job_id=%s
        """, (job_id,))

        job = cursor.fetchone()
        cursor.close()
        conn.close()

        if not job:
            print(f"[DEBUG] Job {job_id} not found in database")
            return jsonify({"success": False, "message": "Job not found"}), 404

        print(f"[DEBUG] Found job: {job.get('job_position')}")

        # Process qualifications for the requirements section
        raw_qualifications = job.get('qualifications') or ""
        requirements = []

        if raw_qualifications:
            if isinstance(raw_qualifications, str):
                requirements = [
                    req.strip() for req in raw_qualifications.replace("\r", "").split("\n")
                    if req.strip()
                ]
                if not requirements:
                    requirements = [
                        req.strip() for req in raw_qualifications.split(",") if req.strip()]
        else:
            requirements = ["No qualifications specified."]

        # Build response payload
        payload = {
            "id": job.get("job_id"),
            "title": job.get("job_position"),
            "job_position": job.get("job_position"),
            "description": job.get("job_description"),
            # Use work_schedule as employment_type
            "employment_type": job.get("work_schedule"),
            "work_schedule": job.get("work_schedule"),
            "requirements": requirements,  # From qualifications
            "qualifications": requirements,
            "min_salary": float(job.get("min_salary", 0)) if job.get("min_salary") else 0,
            "max_salary": float(job.get("max_salary", 0)) if job.get("max_salary") else 0,
            "status": job.get("status"),
            "posted_at": job.get("created_at").isoformat() if job.get("created_at") else None,
            "expiration_date": job.get("job_expiration_date").isoformat() if job.get("job_expiration_date") else None,
            "application_count": job.get("application_count", 0),
            "employer_name": job.get("employer_name"),
            "location": ", ".join(filter(None, [job.get("city"), job.get("province")])),
            "employer_id": job.get("employer_id")
        }

        print(f"[DEBUG] Successfully built payload for job {job_id}")
        return jsonify({"success": True, "job": payload})

    except Exception as e:
        print(f"[DEBUG] ERROR in get_job_details: {str(e)}")
        import traceback
        traceback.print_exc()

        if 'conn' in locals():
            try:
                conn.close()
            except:
                pass

        return jsonify({
            "success": False,
            "message": f"Server error: {str(e)}"
        }), 500


@admin_bp.route("/job/<int:job_id>")
def admin_view_job(job_id):
    """Admin view for job details"""

    conn = create_connection()
    if not conn:
        return "Database connection failed", 500

    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                j.job_id,
                j.job_position,
                j.job_description,
                j.qualifications,
                j.work_schedule,
                j.min_salary,
                j.max_salary,
                j.status,
                j.created_at,
                j.job_expiration_date,
                e.employer_name,
                e.city,
                e.province,
                e.employer_id
            FROM jobs j
            LEFT JOIN employers e ON j.employer_id = e.employer_id
            WHERE j.job_id=%s
        """, (job_id,))

        job = cursor.fetchone()

        if not job:
            return "Job not found", 404

        # Format qualifications as HTML list
        qualifications_html = ""
        if job.get('qualifications'):
            qualifications = [
                q.strip() for q in job['qualifications'].split('\n')
                if q.strip()
            ]
            qualifications_html = "<ul style='margin: 0; padding-left: 20px;'>" + \
                "".join(
                    [f"<li style='margin-bottom: 5px;'>{q}</li>" for q in qualifications]) + "</ul>"
        else:
            qualifications_html = "<p>No qualifications specified.</p>"

        # Format salary
        salary = "Not specified"
        if job.get('min_salary') and job.get('max_salary'):
            salary = f"₱{job['min_salary']:,.2f} - ₱{job['max_salary']:,.2f}"
        elif job.get('min_salary'):
            salary = f"₱{job['min_salary']:,.2f}"
        elif job.get('max_salary'):
            salary = f"₱{job['max_salary']:,.2f}"

        # Format work schedule
        work_schedule = job['work_schedule'].replace(
            '-', ' ').title() if job.get('work_schedule') else "Not specified"

        # Render the job details HTML with your CSS style
        html_content = f"""
        <div class="modal-job-content">
            <div class="job-title-row">
                <h2 style="font-size: 24px; font-weight: bold; color: #2c3e50; margin: 0;">{job['job_position']}</h2>
            </div>
            
            <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                <p style="margin: 8px 0; color: #555;"><strong>Company:</strong> {job['employer_name']}</p>
                <p style="margin: 8px 0; color: #555;"><strong>Location:</strong> {job['city']}, {job['province']}</p>
                <p style="margin: 8px 0; color: #555;"><strong>Salary:</strong> {salary}</p>
                <p style="margin: 8px 0; color: #555;"><strong>Work Schedule:</strong> {work_schedule}</p>
                <p style="margin: 8px 0; color: #555;"><strong>Status:</strong> 
                    <span style="color: {'#27ae60' if job['status'] == 'active' else '#e74c3c'}; font-weight: bold;">
                        {job['status'].title()}
                    </span>
                </p>
                <p style="margin: 8px 0; color: #555;"><strong>Posted:</strong> {job['created_at'].strftime('%B %d, %Y') if job['created_at'] else 'N/A'}</p>
                {f"<p style='margin: 8px 0; color: #555;'><strong>Expires:</strong> {job['job_expiration_date'].strftime('%B %d, %Y')}</p>" if job.get('job_expiration_date') else ''}
            </div>

            <div style="margin-bottom: 20px;">
                <h3 style="font-size: 16px; margin-top: 12px; margin-bottom: 5px; color: #7b1113; font-weight: 600;">Job Description</h3>
                <p style="color: #555; margin-bottom: 25px; font-size: 1rem; line-height: 1.5;">{job['job_description'] or 'No description provided.'}</p>
            </div>

            <div style="margin-bottom: 20px;">
                <h3 style="font-size: 16px; margin-top: 12px; margin-bottom: 5px; color: #7b1113; font-weight: 600;">Requirements</h3>
                {qualifications_html}
            </div>

            <div style="text-align: center; padding-top: 20px; border-top: 1px solid #eee;">
                <a href="/admin/employers/{job['employer_id']}" style="background: #7b1113; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; display: inline-flex; align-items: center; gap: 8px; font-size: 14px;">
                    <i class="fa-solid fa-user-tie"></i> View Employer Profile
                </a>
            </div>
        </div>
        """

        return html_content

    except Exception as e:
        print(f"[ERROR] Failed to load job {job_id}: {e}")
        return f"<p style='color: red; text-align: center; padding: 20px;'>Error loading job details: {str(e)}</p>", 500

    finally:
        if 'conn' in locals():
            conn.close()


def safe_send_email(subject, recipient, body):
    """Send email with proper error handling"""
    if not recipient:
        print(f"❌ No recipient for: {subject}")
        return False

    try:
        msg = Message(
            subject=subject,
            recipients=[recipient],
            html=body,
            sender=("PESO SmartHire", "noreply@pesosmarthire.com")
        )
        mail.send(msg)
        print(f"✅ Email sent to: {recipient}")
        return True
    except Exception as e:
        print(f"❌ Failed to send email to {recipient}: {str(e)}")
        return False


@admin_bp.route("/test-job-report/<int:report_id>")
def test_job_report(report_id):
    """Test if we can fetch report data"""
    conn = create_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT jr.id, jr.job_id, j.job_position, e.employer_name, e.email 
        FROM job_reports jr
        LEFT JOIN jobs j ON jr.job_id = j.job_id
        LEFT JOIN employers e ON j.employer_id = e.employer_id
        WHERE jr.id = %s
    """, (report_id,))

    report = cursor.fetchone()
    cursor.close()
    conn.close()

    return jsonify({
        "success": True,
        "report": report
    })


def ensure_job_report_details_column(cursor):
    cursor.execute("SHOW COLUMNS FROM job_reports LIKE 'details'")
    if not cursor.fetchone():
        cursor.execute(
            "ALTER TABLE job_reports ADD COLUMN details TEXT NULL AFTER reason")


@admin_bp.route("/job_reports/<int:report_id>/action", methods=['POST'])
def handle_job_report_action(report_id):

    data = request.get_json(silent=True) or {}
    action = data.get("action")
    moderator_note = data.get("moderator_note", "").strip()
    days = data.get("days", 0)  # Get the days parameter

    print(
        f"📥 Received request - report_id: {report_id}, action: {action}, days: {days}")

    valid_actions = {"confirm", "reject"}
    if action not in valid_actions:
        return jsonify({"success": False, "message": "Invalid action"}), 400

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT 
                jr.id,
                jr.job_id,
                jr.reason,
                jr.details,
                jr.applicant_id AS reporter_id,
                j.job_position,
                j.employer_id,
                e.employer_name,
                e.email AS employer_email,
                a.email AS reporter_email,
                CONCAT(COALESCE(a.first_name, ''), ' ', COALESCE(a.last_name, '')) AS reporter_name
            FROM job_reports jr
            LEFT JOIN jobs j ON jr.job_id = j.job_id
            LEFT JOIN employers e ON j.employer_id = e.employer_id
            LEFT JOIN applicants a ON jr.applicant_id = a.applicant_id
            WHERE jr.id = %s
        """, (report_id,))
        report = cursor.fetchone()

        if not report:
            print(f"❌ Report {report_id} not found")
            return jsonify({"success": False, "message": "Report not found"}), 404

        job_id = report.get("job_id")
        job_position = report.get("job_position", "Job")
        employer_id = report.get("employer_id")
        reporter_id = report.get("reporter_id")  # The applicant who reported
        report_reason = report.get("reason", "Policy Violation")

        print(
            f"📊 Report found - job_id: {job_id}, job_position: {job_position}")

        if action == "confirm":
            print("🔄 Processing CONFIRM action...")

            # NOTIFY EMPLOYER (The Reported Party)
            if employer_id:
                create_notification(
                    notification_type="report_verdict",
                    title="Job Suspended",
                    message=f"Your job '{job_position}' has been suspended. Reason: {report_reason}. Consequences: Applications cancelled, {days} days to respond.",
                    employer_id=employer_id,
                    related_ids=[job_id]
                )

            if reporter_id:
                create_notification(
                    notification_type="report_verdict",
                    title="Report Confirmed",
                    message=f"Your report against '{job_position}' was confirmed. The job has been suspended.",
                    applicant_id=reporter_id,
                    related_ids=[job_id]
                )

            # Update job status to suspended
            cursor.execute(
                "UPDATE jobs SET status = %s WHERE job_id = %s",
                ("suspended", job_id)
            )
            cursor.execute(
                "UPDATE job_reports SET status = %s, updated_at = NOW() WHERE id = %s",
                ("Confirmed", report_id)
            )
            cursor.execute(
                "UPDATE applications SET status = %s WHERE job_id = %s",
                ("Cancelled", job_id)
            )

            # Get impacted applicants
            cursor.execute("""
                SELECT DISTINCT a.applicant_id, a.email, a.first_name, a.last_name
                FROM applications ap
                JOIN applicants a ON ap.applicant_id = a.applicant_id
                WHERE ap.job_id = %s
            """, (job_id,))
            impacted_applicants = cursor.fetchall() or []

            conn.commit()
            print("✅ Database updates committed")

            # ========== EMAIL SENDING ==========
            email_count = 0

            # 1. Send email to Employer
            employer_email = report.get("employer_email")
            if employer_email:
                try:
                    print(f"📧 Sending email to employer: {employer_email}")
                    safe_send_email(
                        "Job post suspended",
                        employer_email,
                        f"""
                        <p>Hello {report.get('employer_name', 'Employer')},</p>
                        <p>Your job post titled <strong>{job_position}</strong> has been reported and confirmed. 
                        It is now temporarily <strong>suspended</strong> and all applications have been cancelled.</p>
                        <p>You have <strong>{days} days</strong> to respond to this report.</p>
                        <p>Please contact PESO SmartHire admin for more details.</p>
                        """
                    )
                    email_count += 1
                except Exception as e:
                    print(f"❌ Employer email error: {e}")

            # 2. Send emails to Applicants
            for applicant in impacted_applicants:
                applicant_email = applicant.get("email")
                if applicant_email:
                    try:
                        print(
                            f"📧 Sending email to applicant: {applicant_email}")
                        safe_send_email(
                            "Application cancelled",
                            applicant_email,
                            f"""
                            <p>Hi {applicant.get('first_name', 'Applicant')},</p>
                            <p>The job post <strong>{job_position}</strong> was suspended after our investigation. 
                            Your application has been cancelled automatically.</p>
                            <p>We apologize for any inconvenience.</p>
                            """
                        )
                        email_count += 1
                    except Exception as e:
                        print(
                            f"❌ Applicant email error ({applicant['applicant_id']}): {e}")

            # 3. Send email to Reporter
            reporter_email = report.get("reporter_email")
            if reporter_email:
                try:
                    print(f"📧 Sending email to reporter: {reporter_email}")
                    safe_send_email(
                        "Report confirmed",
                        reporter_email,
                        f"""
                        <p>Hi {report.get('reporter_name', 'Applicant')},</p>
                        <p>Your report for <strong>{job_position}</strong> has been confirmed. 
                        The job post is now suspended.</p>
                        <p>Thank you for helping maintain the quality of our platform.</p>
                        """
                    )
                    email_count += 1
                except Exception as e:
                    print(f"❌ Reporter email error: {e}")

            print(f"✅ Process completed - {email_count} emails sent")

            return jsonify({
                "success": True,
                "message": f"Job suspended. Notifications sent to employer and reporter.",
                "job_status": "suspended",
                "report_status": "Confirmed"
            })

        elif action == "reject":
            print("🔄 Processing REJECT action...")
            cursor.execute(
                "UPDATE job_reports SET status = %s, updated_at = NOW() WHERE id = %s",
                ("Rejected", report_id)
            )
            conn.commit()

            # NOTIFY REPORTER (The Applicant)
            if reporter_id:
                verdict_note = moderator_note if moderator_note else "No violation found."
                create_notification(
                    notification_type="report_verdict",
                    title="Report Rejected",
                    message=f"Your report against '{job_position}' was rejected. Verdict: {verdict_note}",
                    applicant_id=reporter_id,
                    related_ids=[job_id]
                )

            # Send rejection email to reporter
            reporter_email = report.get("reporter_email")
            if reporter_email:
                try:
                    print(
                        f"📧 Sending rejection email to reporter: {reporter_email}")
                    safe_send_email(
                        "Report rejected",
                        reporter_email,
                        f"""
                        <p>Hi {report.get('reporter_name', 'Applicant')},</p>
                        <p>Your report for <strong>{job_position}</strong> was rejected. 
                        Our moderators did not find sufficient evidence.</p>
                        {f'<p>Moderator note: {moderator_note}</p>' if moderator_note else ''}
                        <p>Thank you for your understanding.</p>
                        """
                    )
                except Exception as e:
                    print(f"❌ Reporter rejection email error: {e}")

            return jsonify({
                "success": True,
                "message": "Report rejected. Reporter notified.",
                "job_status": None,
                "report_status": "Rejected"
            })

    except Exception as exc:
        conn.rollback()
        print(f"❌ ERROR in handle_job_report_action: {exc}")
        import traceback
        traceback.print_exc()  # This will show the full error traceback
        return jsonify({"success": False, "message": f"Failed to update job status: {str(exc)}"}), 500
    finally:
        cursor.close()
        conn.close()


def ensure_applicant_suspension_column(cursor):
    cursor.execute("SHOW COLUMNS FROM applicants LIKE 'suspension_end_at'")
    if not cursor.fetchone():
        cursor.execute(
            "ALTER TABLE applicants ADD COLUMN suspension_end_at DATETIME NULL AFTER updated_at")


@admin_bp.route("/applicant_reports/<int:report_id>/action", methods=['POST'])
def handle_applicant_report_action(report_id):
    """Moderate reported applicants (confirm/reject) - COMPANY SPECIFIC BLACKLIST"""

    data = request.get_json(silent=True) or {}
    action = data.get("action")
    moderator_note = data.get("moderator_note", "").strip()
    blacklist_days = int(data.get("blacklist_days") or 365)  # Default 1 year

    valid_actions = {"confirm", "reject"}
    if action not in valid_actions:
        return jsonify({"success": False, "message": "Invalid action"}), 400

    conn = create_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT 
                ar.id,
                ar.applicant_id,
                ar.employer_id,
                ar.reason,
                ar.details,
                ar.status AS report_status,
                CONCAT(COALESCE(app.first_name, ''), ' ', COALESCE(app.last_name, '')) AS applicant_name,
                app.email AS applicant_email,
                app.applicant_code,
                emp.employer_name,
                emp.email AS employer_email,
                emp.employer_id
            FROM applicant_reports ar
            LEFT JOIN applicants app ON ar.applicant_id = app.applicant_id
            LEFT JOIN employers emp ON ar.employer_id = emp.employer_id
            WHERE ar.id = %s
        """, (report_id,))
        report = cursor.fetchone()

        if not report:
            return jsonify({"success": False, "message": "Report not found"}), 404

        applicant_id = report.get("applicant_id")
        employer_id = report.get("employer_id")
        employer_name = report.get("employer_name", "the company")

        if action == "confirm":
            # Calculate expiration date
            expires_at = datetime.utcnow() + timedelta(days=blacklist_days)

            # Add to company-specific blacklist
            cursor.execute("""
                INSERT INTO applicant_blacklist 
                (applicant_id, employer_id, reported_by_employer_id, reason, expires_at)
                VALUES (%s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE 
                reason = VALUES(reason), 
                expires_at = VALUES(expires_at),
                blacklisted_at = CURRENT_TIMESTAMP
            """, (
                applicant_id,
                employer_id,
                employer_id,  # The company that reported them
                report.get('reason', ''),
                expires_at
            ))

            # Update any existing applications to this employer to "Blacklisted"
            cursor.execute("""
                UPDATE applications 
                SET status = 'Blacklisted' 
                WHERE applicant_id = %s 
                AND job_id IN (SELECT job_id FROM jobs WHERE employer_id = %s)
            """, (applicant_id, employer_id))

            # Mark report as confirmed
            cursor.execute(
                "UPDATE applicant_reports SET status = %s, updated_at = NOW() WHERE id = %s",
                ("Confirmed", report_id)
            )

            conn.commit()

            report_reason = report.get('reason', 'Violation')

            # NOTIFY APPLICANT (The Reported Party)
            create_notification(
                notification_type="report_verdict",
                title="Account Restricted",
                message=f"You have been restricted from {employer_name}. Reason: {report.get('reason')}. Duration: {blacklist_days} days.",
                applicant_id=applicant_id
            )

            # [Notify EMPLOYER (Reporter)
            create_notification(
                notification_type="report_verdict",
                title="Report Confirmed",
                message=f"Your report against {report.get('applicant_name')} was confirmed. They are now restricted.",
                employer_id=employer_id
            )

            safe_send_email(
                "Application Restrictions - PESO SmartHire",
                report.get("applicant_email"),
                f"""
                <p>Hi {report.get('applicant_name', 'Applicant')},</p>
                <p>A report from <strong>{employer_name}</strong> has been confirmed by our moderation team.</p>
                <p>As a result, you have been <strong>restricted from applying to job posts at {employer_name}</strong>.</p>
                
                <p><strong>What this means:</strong></p>
                <ul>
                    <li>✅ You can still apply to all other companies on our platform</li>
                    <li>✅ You can still login and use all platform features</li>
                    <li>✅ You can still update your profile and resume</li>
                    <li>❌ You cannot apply to {employer_name}'s job posts</li>
                </ul>
                
                <p>This restriction will {'expire automatically after ' + str(blacklist_days) + ' days' if blacklist_days > 0 else 'remain in place until further review'}.</p>
                
                <p>If you believe this is a mistake, please contact PESO SmartHire support.</p>
                """
            )

            # Notify Employer (The Reporter)
            create_notification(
                notification_type="report_verdict",
                title="Report Rejected",
                message=f"Your report against {report.get('applicant_name')} was rejected. Verdict: {verdict_note}",
                employer_id=employer_id
            )

            safe_send_email(
                "Report Confirmed - PESO SmartHire",
                report.get("employer_email"),
                f"""
                <p>Hi {report.get('employer_name', 'Employer')},</p>
                <p>Your report against applicant <strong>{report.get('applicant_name', '')} ({report.get('applicant_code', '')})</strong> has been confirmed.</p>
                <p>This applicant has been restricted from applying to your job posts.</p>
                <p>They can still apply to other employers on our platform.</p>
                <p>The restriction will {'expire after ' + str(blacklist_days) + ' days' if blacklist_days > 0 else 'remain in place until further review'}.</p>
                <p>Thank you for helping maintain platform quality.</p>
                """
            )

            return jsonify({
                "success": True,
                "message": f"Applicant restricted from {employer_name} for {blacklist_days} days.",
                "applicant_status": "Blacklisted from " + employer_name,
                "report_status": "Confirmed"
            })

        # REJECT ACTION (unchanged)
        elif action == "reject":
            cursor.execute(
                "UPDATE applicant_reports SET status = %s, updated_at = NOW() WHERE id = %s",
                ("Rejected", report_id)
            )
            conn.commit()

            verdict_note = moderator_note if moderator_note else "Insufficient evidence."

            # NOTIFY EMPLOYER (The Reporter)
            create_notification(
                notification_type="report_verdict",
                title="Report Rejected",
                message=f"Your report against {report.get('applicant_name')} was rejected. Verdict: {verdict_note}",
                employer_id=employer_id
            )

            safe_send_email(
                "Report Rejected - PESO SmartHire",
                report.get("employer_email"),
                f"""
                <p>Hi {report.get('employer_name', 'Employer')},</p>
                <p>Your report against {report.get('applicant_name', 'an applicant')} was reviewed but we did not find sufficient evidence of violation.</p>
                {f'<p><strong>Moderator note:</strong> {moderator_note}</p>' if moderator_note else ''}
                <p>The applicant continues to have full access to the platform.</p>
                <p>Thank you for your understanding.</p>
                """
            )

            return jsonify({
                "success": True,
                "message": "Report rejected and employer notified.",
                "applicant_status": None,
                "report_status": "Rejected"
            })

    except Exception as exc:
        conn.rollback()
        print(f"[v2] Failed to update applicant report {report_id}: {exc}")
        return jsonify({"success": False, "message": "Failed to process report."}), 500
    finally:
        cursor.close()
        conn.close()


def is_applicant_blacklisted(applicant_id, employer_id):
    """Check if applicant is blacklisted from a specific employer"""
    conn = create_connection()
    if not conn:
        return False

    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT COUNT(*) FROM applicant_blacklist 
            WHERE applicant_id = %s AND employer_id = %s 
            AND (expires_at IS NULL OR expires_at > NOW())
        """, (applicant_id, employer_id))

        is_blacklisted = cursor.fetchone()[0] > 0
        return is_blacklisted
    except Exception as e:
        print(f"Error checking blacklist: {e}")
        return False
    finally:
        cursor.close()
        conn.close()


@admin_bp.route('/update_report_status', methods=['POST'])
def update_report_status():

    report_id = request.form.get("report_id")
    new_status = request.form.get("status")
    days = request.form.get("days", 0)

    print(f"🔄 Updating report {report_id} to {new_status} with {days} days")

    conn = create_connection()
    if not conn:
        return jsonify({"status": "error", "message": "Database connection failed"})

    cursor = conn.cursor(dictionary=True)

    try:
        # 1. Update the report status in the database
        cursor.execute("""
            UPDATE job_reports
            SET status = %s
            WHERE id = %s
        """, (new_status, report_id))

        # 2. Fetch report details (Crucial for identifying who to notify)
        cursor.execute("""
            SELECT 
                jr.job_id,
                jr.applicant_id AS reporter_id,
                jr.reason,
                j.job_position,
                j.employer_id,
                e.employer_name,
                e.email AS employer_email,
                a.email AS reporter_email,
                CONCAT(a.first_name, ' ', a.last_name) AS reporter_name
            FROM job_reports jr
            LEFT JOIN jobs j ON jr.job_id = j.job_id
            LEFT JOIN employers e ON j.employer_id = e.employer_id
            LEFT JOIN applicants a ON jr.applicant_id = a.applicant_id
            WHERE jr.id = %s
        """, (report_id,))
        report = cursor.fetchone()

        if report:
            job_id = report.get('job_id')
            job_position = report.get('job_position') or "Job Post"
            employer_id = report.get('employer_id')
            reporter_id = report.get('reporter_id')
            report_reason = report.get('reason', 'Policy Violation')

            # === CASE A: CONFIRMED ===
            if new_status == "Confirmed":
                # [Existing] Update jobs and applications status
                cursor.execute(
                    "UPDATE jobs SET status = 'suspended' WHERE job_id = %s", (job_id,))
                cursor.execute(
                    "UPDATE applications SET status = 'Cancelled' WHERE job_id = %s", (job_id,))

                # [Existing] Get impacted applicants for email
                cursor.execute("""
                    SELECT DISTINCT a.applicant_id, a.email, a.first_name
                    FROM applications ap
                    JOIN applicants a ON ap.applicant_id = a.applicant_id
                    WHERE ap.job_id = %s AND a.email IS NOT NULL
                """, (job_id,))
                impacted_applicants = cursor.fetchall()

                # 1. Notify Employer (Using 'report_verdict' type)
                if employer_id:
                    create_notification(
                        notification_type="report_verdict",
                        title="Job Suspended",
                        message=f"Your job '{job_position}' has been suspended. Reason: {report_reason}.",
                        employer_id=employer_id,
                        related_ids=[job_id]
                    )

                # 2. Notify Reporter
                if reporter_id:
                    create_notification(
                        notification_type="report_verdict",
                        title="Report Confirmed",
                        message=f"Your report on '{job_position}' was confirmed. Action taken.",
                        applicant_id=reporter_id,
                        related_ids=[job_id]
                    )

                if report.get('employer_email'):
                    safe_send_email(
                        "Job Post Suspended", report['employer_email'], f"<p>Your job '{job_position}' was suspended.</p>")

                if report.get('reporter_email'):
                    safe_send_email(
                        "Report Confirmed", report['reporter_email'], f"<p>Your report on '{job_position}' was confirmed.</p>")

                for app_user in impacted_applicants:
                    if app_user.get('email'):
                        safe_send_email(
                            "Application Cancelled", app_user['email'], f"<p>Your application to '{job_position}' was cancelled.</p>")

            # === CASE B: REJECTED ===
            elif new_status == "Rejected":
                # --- FIX: SEND IN-APP NOTIFICATION ---
                if reporter_id:
                    create_notification(
                        notification_type="report_verdict",
                        title="Report Rejected",
                        message=f"Your report on '{job_position}' was rejected. No violation found.",
                        applicant_id=reporter_id,
                        related_ids=[job_id]
                    )

                # [Existing Email Logic]
                if report.get('reporter_email'):
                    safe_send_email(
                        "Report Rejected",
                        report['reporter_email'],
                        f"<p>Your report for <strong>{job_position}</strong> was reviewed and rejected. We did not find sufficient evidence of a violation.</p>"
                    )

        conn.commit()
        return jsonify({"status": "success", "message": "Status updated and notifications sent!"})

    except Exception as e:
        conn.rollback()
        print(f"❌ Database error: {e}")
        return jsonify({"status": "error", "message": f"Database error: {str(e)}"})


@admin_bp.route("/applicants/<int:applicant_id>")
def view_applicant(applicant_id):
    conn = create_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT * FROM applicants WHERE applicant_id = %s", (applicant_id,)
    )
    applicant = cursor.fetchone()
    cursor.close()
    conn.close()

    if not applicant:
        flash("Applicant not found", "danger")
        return redirect(url_for("admin.applicants_management"))

    # Only prepare documents if applicant is non-Lipeño
    documents = []
    if not applicant.get("is_from_lipa") and applicant.get("recommendation_letter_path"):
        documents.append({
            "name": "Recommendation Letter",
            "last_updated": applicant.get("recommendation_letter_uploaded_at") or applicant.get("updated_at"),
            "expires_at": applicant.get("recommendation_letter_expiry")
        })

    referrer = request.referrer
    from_notifications = False
    if referrer and "/admin/notifications" in referrer:
        from_notifications = True

    return render_template(
        "Admin/applicant_profile.html",
        applicant=applicant,
        from_notifications=from_notifications,
        documents=documents
    )


# ===== Admin: Employers Management =====
@admin_bp.route("/employers")
def employers_management():
    return render_template("Admin/admin_employer.html")


@admin_bp.route("/employers/for-approval")
def employers_for_approval():
    """Show employers (local and international) needing approval"""

    conn = create_connection()
    cursor = conn.cursor(dictionary=True)

    # Get employers with Pending status
    cursor.execute("""
        SELECT employer_id, employer_name, recruitment_type,
               created_at, status
        FROM employers
        WHERE status = 'Pending'
        ORDER BY created_at DESC
    """)
    employers = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("Admin/employers_for_approval.html", employers=employers)


@admin_bp.route("/employers/view-all")
def employers_view_all():
    """Show all employers with Local/International filter"""

    conn = create_connection()
    cursor = conn.cursor(dictionary=True)

    # Get all employers (all statuses)
    cursor.execute("""
        SELECT employer_id, employer_name, recruitment_type,
               created_at, status
        FROM employers
        ORDER BY created_at DESC
    """)
    employers = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("Admin/employers_view_all.html", employers=employers)


@admin_bp.route("/employers/<int:employer_id>")
def view_employer(employer_id):
    conn = create_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT * FROM employers WHERE employer_id = %s", (employer_id,)
    )
    employer = cursor.fetchone()
    cursor.close()
    conn.close()

    if not employer:
        flash("Employer not found", "danger")
        return redirect(url_for("admin.employers_management"))

    LOCAL_DOCUMENTS = {
        "business_permit_path": "business_permit_expiry",
        "philiobnet_registration_path": "philiobnet_registration_expiry",
        "job_orders_of_client_path": "job_orders_expiry",
        "dole_no_pending_case_path": "dole_no_pending_case_expiry",
        "dole_authority_to_recruit_path": "dole_authority_expiry",
    }

    INTERNATIONAL_DOCUMENTS = {
        "business_permit_path": "business_permit_expiry",
        "philiobnet_registration_path": "philiobnet_registration_expiry",
        "job_orders_of_client_path": "job_orders_expiry",
        "dmw_no_pending_case_path": "dmw_no_pending_case_expiry",
        "license_to_recruit_path": "license_to_recruit_expiry",
    }

    UPLOADED_AT_MAP = {
        "business_permit_path": "business_permit_uploaded_at",
        "philiobnet_registration_path": "philiobnet_uploaded_at",
        "job_orders_of_client_path": "job_orders_uploaded_at",
        "dole_no_pending_case_path": "dole_no_pending_uploaded_at",
        "dole_authority_to_recruit_path": "dole_authority_uploaded_at",
        "dmw_no_pending_case_path": "dmw_no_pending_uploaded_at",
        "license_to_recruit_path": "license_to_recruit_uploaded_at",
    }

    ALL_DOCUMENTS = {**LOCAL_DOCUMENTS, **INTERNATIONAL_DOCUMENTS}

    documents = []
    for file_field, expiry_field in ALL_DOCUMENTS.items():
        if employer.get(file_field):  # Only process if file exists
            uploaded_at_field = UPLOADED_AT_MAP.get(file_field)
            last_updated_value = employer.get(uploaded_at_field)

            expires_at_value = employer.get(expiry_field)
            if expires_at_value and isinstance(expires_at_value, str):
                expires_at_value = datetime.fromisoformat(expires_at_value)

            documents.append({
                "name": file_field.replace("_path", "").replace("_", " ").title(),
                "last_updated": last_updated_value,
                "expires_at": expires_at_value
            })

    referrer = request.referrer
    from_notifications = False
    if referrer and "/admin/notifications" in referrer:
        from_notifications = True

    return render_template(
        "Admin/employer_profile.html",
        employer=employer,
        documents=documents,
        from_notifications=from_notifications,
        recruitment_type_change_pending=employer.get(
            "recruitment_type_change_pending", 0)
    )


@admin_bp.route("/delete-rejected-employer/<int:employer_id>", methods=["POST"])
def delete_rejected_employer(employer_id):
    """Delete a rejected employer record from the system.

    This allows the employer to retry registration with the same email/info.
    All associated documents and files are also deleted.
    """
    try:
        data = request.get_json() if request.is_json else {}

        conn = create_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM employers WHERE employer_id = %s AND status = 'Rejected'",
            (employer_id,)
        )
        employer = cursor.fetchone()

        if not employer:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Employer not found or not in Rejected status"}), 404

        file_fields = [
            "company_logo_path",
            "business_permit_path",
            "philiobnet_registration_path",
            "job_orders_of_client_path",
            "dole_no_pending_case_path",
            "dole_authority_to_recruit_path",
            "dmw_no_pending_case_path",
            "license_to_recruit_path"
        ]

        for field in file_fields:
            file_path = employer.get(field)
            if file_path:
                try:
                    full_path = os.path.join("static", file_path)
                    if os.path.exists(full_path):
                        os.remove(full_path)
                        logger.info(f"Deleted file: {full_path}")
                except Exception as e:
                    logger.warning(f"Failed to delete file {file_path}: {e}")

        # Delete the employer record from database
        cursor.execute(
            "DELETE FROM employers WHERE employer_id = %s",
            (employer_id,)
        )
        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({
            "success": True,
            "message": f"Rejected employer record deleted. They can now re-register with the same information."
        })

    except Exception as e:
        logger.exception(f"Error deleting rejected employer: {e}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "message": f"An error occurred: {str(e)}"}), 500


@admin_bp.route("/update_local_employer_status/<int:employer_id>", methods=["POST"])
def update_local_employer_status(employer_id):
    try:
        data = request.get_json()
        print(f"[v1] Received data for employer {employer_id}: {data}")

        if not data or "action" not in data:
            print("[v1] No action provided in request")
            return jsonify({"success": False, "message": "No action provided."}), 400

        action = data["action"]
        reason = None
        documents_to_reupload = None

        conn = create_connection()
        if not conn:
            print("[v1] Database connection failed")
            return jsonify({"success": False, "message": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM employers WHERE employer_id = %s", (employer_id,))
        employer = cursor.fetchone()

        if not employer:
            print(f"[v1] Employer {employer_id} not found in DB")
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Employer not found"}), 404

        is_new_employer = employer.get("must_change_password") == 1

        temp_password_plain = None

        if action == "approved":
            if is_new_employer:
                # Only generate new credentials for first-time approval
                temp_password_plain = secrets.token_urlsafe(8)
                password_hash = generate_password_hash(temp_password_plain)
                cursor.execute(
                    "UPDATE employers SET password_hash = %s, temp_password = %s, must_change_password = 1 WHERE employer_id = %s",
                    (password_hash, temp_password_plain, employer_id)
                )
            else:
                # EXISTING employer - keep current credentials, DO NOT reset must_change_password
                temp_password_plain = employer.get("temp_password")

            new_status = "Approved"
            subject = "PESO SmartHire - Local Recruitment Account Approved"

            if is_new_employer:
                credentials_block = f"""
                <p>Included below are your login credentials:</p>
                <ul>
                    <li>Employer Code: {employer['employer_code']}</li>
                    <li>Email: {employer['email']}</li>
                    <li>Phone Number: {employer['phone']}</li>
                    <li>Password: {temp_password_plain}</li>
                </ul>
                <p><strong>You are required to change your password upon logging in for security purposes.</strong></p>
                """
            else:
                credentials_block = """
                <p><strong>Use your existing login credentials to access your account.</strong></p>
                """

            body = f"""
            <p>Dear {employer['employer_name']},</p>
            <p>This is PESO SmartHire Team.</p>
            <p>Congratulations! Your local recruitment account has been reviewed and approved!</p>
            <p>You may now post job orders and access your employer dashboard to manage your recruitment activities.</p>
            {credentials_block}
            <p>To get started, visit our platform and log in with your credentials. You can then begin posting job orders and managing your recruitment needs.</p>
            <p>If you have any questions or need assistance, please don't hesitate to contact our support team.</p>
            <p>Thank you for partnering with PESO SmartHire!</p>
            <p>— PESO SmartHire Admin</p>
            """
            success_message = "Local employer approved successfully! Email notification sent."

        elif action == "rejected":
            new_status = "Rejected"
            reason = data.get("reason") if isinstance(data, dict) else None
            subject = "PESO SmartHire - Local Recruitment Account Status Update"
            reason_block = f"<p><strong>Reason:</strong> {reason}</p>" if reason else ""
            body = f"""
            <p>Dear {employer['employer_name']},</p>
            <p>This is PESO SmartHire Team.</p>
            <p>We regret to inform you that your local recruitment account application has been reviewed but did not meet the current requirements.</p>
            {reason_block}
            <p>Please review the requirements and feel free to reapply in the future once you have met all the necessary qualifications.</p>
            <p>If you have any questions regarding this decision, please contact our support team.</p>
            <p>Thank you for your interest in PESO SmartHire.</p>
            <p>— PESO SmartHire Admin</p>
            """

            is_new_registration = (employer['status'] == 'Pending') and (
                employer.get("recruitment_type_change_pending", 0) == 0)

            if is_new_registration:
                success_message = "Employer application rejected and record deleted."

                try:
                    msg = Message(subject=subject, recipients=[
                                  employer["email"]], html=body)
                    mail.send(msg)
                except Exception as email_error:
                    logger.error(
                        f"Failed to send rejection email: {email_error}")
                    # Even if email fails, we still want to delete the record

                file_fields = [
                    "company_logo_path",
                    "business_permit_path",
                    "philiobnet_registration_path",
                    "job_orders_of_client_path",
                    "dole_no_pending_case_path",
                    "dole_authority_to_recruit_path",
                    "dmw_no_pending_case_path",
                    "license_to_recruit_path"
                ]

                for field in file_fields:
                    file_path = employer.get(field)
                    if file_path:
                        try:
                            full_path = os.path.join("static", file_path)
                            if os.path.exists(full_path):
                                os.remove(full_path)
                                logger.info(f"Deleted file: {full_path}")
                        except Exception as e:
                            logger.warning(
                                f"Failed to delete file {file_path}: {e}")

                # Delete the record
                cursor.execute(
                    "DELETE FROM employers WHERE employer_id = %s",
                    (employer_id,)
                )
                conn.commit()
                cursor.close()
                conn.close()

                return jsonify({"success": True, "message": success_message})
            else:
                success_message = "Local employer rejected. Notification email sent."

        elif action == "reupload":
            # Logic from your request
            if not employer.get("temp_password"):
                temp_password_plain = secrets.token_urlsafe(8)
                password_hash = generate_password_hash(temp_password_plain)

                cursor.execute(
                    "UPDATE employers SET password_hash = %s, temp_password = %s WHERE employer_id = %s",
                    (password_hash, temp_password_plain, employer_id)
                )
            else:
                temp_password_plain = employer["temp_password"]

            new_status = "Reupload"

            requested = data.get("document_name")
            if isinstance(requested, list):
                requested_list = requested
            elif isinstance(requested, str) and requested:
                requested_list = [requested]
            else:
                requested_list = []

            # Normalize names so they match DB field prefixes
            normalized_map = {
                "Business Permit": "business_permit",
                "PhilJobNet Registration": "philiobnet_registration",
                "Job Orders of Client": "job_orders_of_client",
                "DOLE - No Pending Case Certificate": "dole_no_pending_case",
                "DOLE - Authority to Recruit": "dole_authority_to_recruit",
                "DMW - No Pending Case Certificate": "dmw_no_pending_case",
                "DMW - License to Recruit": "license_to_recruit",
                "Company Logo": "company_logo"
            }

            normalized_docs = [
                normalized_map.get(doc.strip(), doc.strip(
                ).lower().replace(' ', '_').replace('-', '_'))
                for doc in requested_list
            ]

            # Save normalized field names in JSON column
            documents_to_reupload = json.dumps(
                normalized_docs) if normalized_docs else None

            docs_block = ""
            if requested_list:
                docs_html = "".join([f"<li>{d}</li>" for d in requested_list])
                docs_block = f"<p>The documents we specifically request you to re-upload are:</p><ul>{docs_html}</ul>"

            subject = "PESO SmartHire - Local Recruitment Documents Update Required"
            body = f"""
            <p>Dear {employer['employer_name']},</p>
            <p>This is PESO SmartHire Team.</p>
            <p>We have reviewed your local recruitment account and noticed that some of your required documents need to be updated or are missing important information.</p>
            {docs_block}
            <p>Please log in to your account and re-upload the required documents through your employer dashboard as soon as possible.</p>
            <p>Here are your login credentials:</p>
            <ul>
                <li>Employer ID: {employer['employer_code']}</li>
                <li>Email: {employer['email']}</li>
                <li>Phone Number: {employer['phone']}</li>
                <li>Password: {temp_password_plain}</li>
            </ul>
            <p><strong>Please change your password after logging in for security purposes.</strong></p>
            <p>Once you have updated your documents, we will review them promptly and notify you of the status.</p>
            <p>If you need any assistance, please contact our support team.</p>
            <p>Thank you for your cooperation!</p>
            <p>— PESO SmartHire Admin</p>
            """
            success_message = "Re-upload request sent. Email notification with login credentials sent to local employer."

        else:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Invalid action."}), 400

        # Update employer status and optional rejection reason
        if new_status == "Approved":
            is_active_value = 1
        else:
            is_active_value = 0

        if reason:
            cursor.execute(
                "UPDATE employers SET status = %s, rejection_reason = %s, is_active = %s, documents_to_reupload = %s WHERE employer_id = %s",
                (new_status, reason, is_active_value,
                 documents_to_reupload, employer_id)
            )
        else:
            cursor.execute(
                "UPDATE employers SET status = %s, is_active = %s, documents_to_reupload = %s, approved_at = NOW() WHERE employer_id = %s",
                (new_status, is_active_value, documents_to_reupload, employer_id)
            )
        conn.commit()

        try:
            msg = Message(subject=subject, recipients=[
                          employer["email"]], html=body)
            mail.send(msg)
        except Exception as e:
            print(f"Failed to send email: {e}")

        cursor.close()
        conn.close()

        return jsonify({"success": True, "message": success_message})

    except Exception as e:
        print(f"[v1] Error updating status or sending email: {e}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "message": f"An error occurred: {str(e)}"}), 500


@admin_bp.route("/update_international_employer_status/<int:employer_id>", methods=["POST"])
def update_international_employer_status(employer_id):
    try:
        data = request.get_json()
        print(f"[v1] Received data for employer {employer_id}: {data}")

        if not data or "action" not in data:
            print("[v1] No action provided in request")
            return jsonify({"success": False, "message": "No action provided."}), 400

        action = data["action"]
        reason = None
        documents_to_reupload = None

        conn = create_connection()
        if not conn:
            print("[v1] Database connection failed")
            return jsonify({"success": False, "message": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM employers WHERE employer_id = %s", (employer_id,))
        employer = cursor.fetchone()

        if not employer:
            print(f"[v1] Employer {employer_id} not found in DB")
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Employer not found"}), 404

        is_new_employer = employer.get("must_change_password") == 1

        temp_password_plain = None

        if action == "approved":
            if is_new_employer:
                # Only generate new credentials for first-time approval
                temp_password_plain = secrets.token_urlsafe(8)
                password_hash = generate_password_hash(temp_password_plain)
                cursor.execute(
                    "UPDATE employers SET password_hash = %s, temp_password = %s, must_change_password = 1 WHERE employer_id = %s",
                    (password_hash, temp_password_plain, employer_id)
                )
            else:
                # EXISTING employer - keep current credentials, DO NOT reset must_change_password
                temp_password_plain = employer.get("temp_password")

            new_status = "Approved"
            subject = "PESO SmartHire - International Recruitment Account Approved"

            if is_new_employer:
                credentials_block = f"""
                <p>Included below are your login credentials:</p>
                <ul>
                    <li>Employer Code: {employer['employer_code']}</li>
                    <li>Email: {employer['email']}</li>
                    <li>Phone Number: {employer['phone']}</li>
                    <li>Password: {temp_password_plain}</li>
                </ul>
                <p><strong>You are required to change your password upon logging in for security purposes.</strong></p>
                """
            else:
                credentials_block = """
                <p><strong>Use your existing login credentials to access your account.</strong></p>
                """

            body = f"""
            <p>Dear {employer['employer_name']},</p>
            <p>This is PESO SmartHire Team.</p>
            <p>Congratulations! Your international recruitment account has been reviewed and approved!</p>
            <p>You may now post job orders and access your employer dashboard to manage your recruitment activities.</p>
            {credentials_block}
            <p>To get started, visit our platform and log in with your credentials. You can then begin posting job orders and managing your recruitment needs.</p>
            <p>If you have any questions or need assistance, please don't hesitate to contact our support team.</p>
            <p>Thank you for partnering with PESO SmartHire!</p>
            <p>— PESO SmartHire Admin</p>
            """
            success_message = "International employer approved successfully! Email notification sent."

        elif action == "rejected":
            new_status = "Rejected"
            reason = data.get("reason") if isinstance(data, dict) else None
            subject = "PESO SmartHire - International Recruitment Account Status Update"
            reason_block = f"<p><strong>Reason:</strong> {reason}</p>" if reason else ""
            body = f"""
            <p>Dear {employer['employer_name']},</p>
            <p>This is PESO SmartHire Team.</p>
            <p>We regret to inform you that your international recruitment account application has been reviewed but did not meet the current requirements.</p>
            {reason_block}
            <p>Please review the requirements and feel free to reapply in the future once you have met all the necessary qualifications.</p>
            <p>If you have any questions regarding this decision, please contact our support team.</p>
            <p>Thank you for your interest in PESO SmartHire.</p>
            <p>— PESO SmartHire Admin</p>
            """

            is_new_registration = (employer['status'] == 'Pending') and (
                employer.get("recruitment_type_change_pending", 0) == 0)

            if is_new_registration:
                success_message = "Employer application rejected and record deleted."

                try:
                    msg = Message(subject=subject, recipients=[
                                  employer["email"]], html=body)
                    mail.send(msg)
                except Exception as email_error:
                    logger.error(
                        f"Failed to send rejection email: {email_error}")
                    # Even if email fails, we still want to delete the record

                file_fields = [
                    "company_logo_path",
                    "business_permit_path",
                    "philiobnet_registration_path",
                    "job_orders_of_client_path",
                    "dole_no_pending_case_path",
                    "dole_authority_to_recruit_path",
                    "dmw_no_pending_case_path",
                    "license_to_recruit_path"
                ]

                for field in file_fields:
                    file_path = employer.get(field)
                    if file_path:
                        try:
                            full_path = os.path.join("static", file_path)
                            if os.path.exists(full_path):
                                os.remove(full_path)
                                logger.info(f"Deleted file: {full_path}")
                        except Exception as e:
                            logger.warning(
                                f"Failed to delete file {file_path}: {e}")

                # Delete the record
                cursor.execute(
                    "DELETE FROM employers WHERE employer_id = %s",
                    (employer_id,)
                )
                conn.commit()
                cursor.close()
                conn.close()

                return jsonify({"success": True, "message": success_message})
            else:
                success_message = "International employer rejected. Notification email sent."

        elif action == "reupload":
            if not employer.get("temp_password"):
                temp_password_plain = secrets.token_urlsafe(8)
                password_hash = generate_password_hash(temp_password_plain)

                cursor.execute(
                    "UPDATE employers SET password_hash = %s, temp_password = %s WHERE employer_id = %s",
                    (password_hash, temp_password_plain, employer_id)
                )
            else:
                temp_password_plain = employer["temp_password"]

            new_status = "Reupload"

            requested = data.get("document_name")
            if isinstance(requested, list):
                requested_list = requested
            elif isinstance(requested, str) and requested:
                requested_list = [requested]
            else:
                requested_list = []

            normalized_map = {
                "Business Permit": "business_permit",
                "PhilJobNet Registration": "philiobnet_registration",
                "Job Orders of Client": "job_orders_of_client",
                "DOLE - No Pending Case Certificate": "dole_no_pending_case",
                "DOLE - Authority to Recruit": "dole_authority_to_recruit",
                "DMW - No Pending Case Certificate": "dmw_no_pending_case",
                "DMW - License to Recruit": "license_to_recruit",
                "Company Logo": "company_logo"
            }

            normalized_docs = [
                normalized_map.get(doc.strip(), doc.strip(
                ).lower().replace(' ', '_').replace('-', '_'))
                for doc in requested_list
            ]

            documents_to_reupload = json.dumps(
                normalized_docs) if normalized_docs else None

            docs_block = ""
            if requested_list:
                docs_html = "".join([f"<li>{d}</li>" for d in requested_list])
                docs_block = f"<p>The documents we specifically request you to re-upload are:</p><ul>{docs_html}</ul>"

            subject = "PESO SmartHire - International Recruitment Documents Update Required"
            body = f"""
            <p>Dear {employer['employer_name']},</p>
            <p>This is PESO SmartHire Team.</p>
            <p>We have reviewed your international recruitment account and noticed that some of your required documents need to be updated or are missing important information.</p>
            {docs_block}
            <p>Please log in to your account and re-upload the required documents through your employer dashboard as soon as possible.</p>
            <p>Here are your login credentials:</p>
            <ul>
                <li>Employer ID: {employer['employer_code']}</li>
                <li>Email: {employer['email']}</li>
                <li>Phone Number: {employer['phone']}</li>
                <li>Password: {temp_password_plain}</li>
            </ul>
            <p><strong>Please change your password after logging in for security purposes.</strong></p>
            <p>Once you have updated your documents, we will review them promptly and notify you of the status.</p>
            <p>If you need any assistance, please contact our support team.</p>
            <p>Thank you for your cooperation!</p>
            <p>— PESO SmartHire Admin</p>
            """
            success_message = "Re-upload request sent. Email notification with login credentials sent to international employer."

        else:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Invalid action."}), 400

        # Update employer status and save which documents need reupload
        if new_status == "Approved":
            is_active_value = 1
        else:
            is_active_value = 0

        if reason:
            cursor.execute(
                "UPDATE employers SET status = %s, rejection_reason = %s, is_active = %s, documents_to_reupload = %s WHERE employer_id = %s",
                (new_status, reason, is_active_value,
                 documents_to_reupload, employer_id)
            )
        else:
            cursor.execute(
                "UPDATE employers SET status = %s, is_active = %s, documents_to_reupload = %s, approved_at = NOW() WHERE employer_id = %s",
                (new_status, is_active_value, documents_to_reupload, employer_id)
            )
        conn.commit()

        try:
            msg = Message(subject=subject, recipients=[
                          employer["email"]], html=body)
            mail.send(msg)
        except Exception as e:
            print(f"Failed to send email: {e}")

        cursor.close()
        conn.close()

        return jsonify({"success": True, "message": success_message})

    except Exception as e:
        print(f"[v1] Error updating status or sending email: {e}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "message": f"An error occurred: {str(e)}"}), 500


@admin_bp.route("/reupload-recruitment-type-change/<int:employer_id>", methods=["POST"])
def reupload_recruitment_type_change(employer_id):
    """
    Admin requests employer to reupload documents for new recruitment type
    Shows only DOLE or DMW documents based on new recruitment type
    """
    try:
        data = request.get_json()

        conn = create_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM employers WHERE employer_id = %s",
            (employer_id,)
        )
        employer = cursor.fetchone()

        if not employer:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Employer not found"}), 404

        # NEW employer: must_change_password = 1 (hasn't changed password yet)
        is_new_employer = employer.get("must_change_password") == 1

        new_recruitment_type = employer["recruitment_type"]
        # Build document list based on NEW recruitment type
        if new_recruitment_type == "Local":
            documents_needed = ["dole_no_pending_case",
                                "dole_authority_to_recruit"]
            doc_labels = ["DOLE - No Pending Case Certificate",
                          "DOLE - Authority to Recruit"]
        else:  # International
            documents_needed = ["dmw_no_pending_case", "license_to_recruit"]
            doc_labels = ["DMW - No Pending Case Certificate",
                          "DMW - License to Recruit"]

        documents_to_reupload_json = json.dumps(documents_needed)

        # Update employer: set to Reupload status, mark as restricted to documents tab
        cursor.execute("""
            UPDATE employers
            SET status = %s,
                is_active = 0,
                recruitment_type_change_pending = 1,
                documents_to_reupload = %s
            WHERE employer_id = %s
        """, ("Reupload", documents_to_reupload_json, employer_id))

        conn.commit()

        temp_password_plain = None
        # If new and missing temp_password, generate one. Otherwise keep existing temp_password (DO NOT reset for existing)
        if is_new_employer and not employer.get("temp_password"):
            temp_password_plain = secrets.token_urlsafe(8)
            password_hash = generate_password_hash(temp_password_plain)

            cursor.execute(
                "UPDATE employers SET password_hash = %s, temp_password = %s WHERE employer_id = %s",
                (password_hash, temp_password_plain, employer_id)
            )
            conn.commit()
        else:
            temp_password_plain = employer.get("temp_password")

        doc_list = "".join([f"<li>{label}</li>" for label in doc_labels])

        if is_new_employer:
            credentials_block = f"""
            <p>Please log in to your account using the credentials below and upload these documents in the <strong>Documents tab only</strong> (other features are temporarily restricted).</p>
            <ul>
                <li>Employer Code: {employer['employer_code']}</li>
                <li>Email: {employer['email']}</li>
                <li>Phone Number: {employer['phone']}</li>
                <li>Password: {temp_password_plain}</li>
            </ul>
            """
        else:
            credentials_block = f"""
            <p>Please log in to your account using your existing credentials and upload these documents in the <strong>Documents tab only</strong> (other features are temporarily restricted).</p>
            <ul>
                <li>Employer Code: {employer['employer_code']}</li>
                <li>Email: {employer['email']}</li>
            </ul>
            """

        subject = f"PESO SmartHire - Reupload Required for {new_recruitment_type} Recruitment"
        body = f"""
        <p>Dear {employer['employer_name']},</p>
        <p>This is PESO SmartHire Team.</p>
        <p>We have reviewed your recruitment type change to <strong>{new_recruitment_type} recruitment</strong> and need you to reupload the required documents.</p>
        <p><strong>You must reupload the following documents:</strong></p>
        <ul>{doc_list}</ul>
        {credentials_block}
        <p>Once you have uploaded the required documents, we will review them and notify you of the status.</p>
        <p>Thank you for your cooperation!</p>
        <p>— PESO SmartHire Admin</p>
        """

        msg = Message(subject=subject, recipients=[
                      employer["email"]], html=body)
        mail.send(msg)

        cursor.close()
        conn.close()

        return jsonify({"success": True, "message": "Reupload request sent to employer"})

    except Exception as e:
        print(f"[v0] Error requesting recruitment type reupload: {e}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route("/account-settings", methods=["GET", "POST"])
def account_settings():

    try:
        conn = create_connection()
        if not conn:
            flash("Database connection failed", "danger")
            return redirect(url_for("admin.admin_home"))

        if request.method == "POST":
            new_email = request.form.get("email")

            update_query = "UPDATE admin SET email = %s WHERE admin_id = %s"
            result = run_query(conn, update_query,
                               (new_email, session["admin_id"]))

            if result:
                flash("Email updated successfully", "success")
                session["admin_email"] = new_email
            else:
                flash("Failed to update email", "danger")

            conn.close()
            return redirect(url_for("admin.account_settings"))

        # --- GET Request ---
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT admin_code, email FROM admin WHERE admin_id = %s",
                       (session["admin_id"],))
        admin = cursor.fetchone()
        cursor.close()
        conn.close()

        if not admin:
            flash("Admin not found", "danger")
            return redirect(url_for("admin.admin_home"))

        return render_template("Admin/admin_acc.html", admin=admin)

    except Exception as e:
        print(f"[account_settings] Error: {e}")
        flash("An error occurred while loading account settings", "danger")
        return redirect(url_for("admin.admin_home"))


# This is for recruitment changing on edit state
@admin_bp.route("/approve-recruitment-type-change/<int:employer_id>", methods=["POST"])
def approve_recruitment_type_change(employer_id):
    """
    Admin approves employer recruitment type change
    Sets status to Approved and sends email with existing login credentials
    """
    try:
        conn = create_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM employers WHERE employer_id = %s",
            (employer_id,)
        )
        employer = cursor.fetchone()

        if not employer:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Employer not found"}), 404

        # Update status to Approved
        cursor.execute("""
            UPDATE employers
            SET status = %s,
                recruitment_type_change_pending = 0,
                is_active = 1,
                approved_at = NOW()
            WHERE employer_id = %s
        """, ("Approved", employer_id))

        conn.commit()

        # Send approval email with EXISTING login credentials
        subject = f"PESO SmartHire - {employer['recruitment_type']} Recruitment Approved"
        body = f"""
        <p>Dear {employer['employer_name']},</p>
        <p>This is PESO SmartHire Team.</p>
        <p>Congratulations! Your recruitment type change has been reviewed and approved.</p>
        <p>You may now use your account to manage {employer['recruitment_type'].lower()} recruitment activities.</p>
        <p><strong>To log in, use your existing credentials:</strong></p>
        <ul>
            <li>Employer Code: {employer['employer_code']}</li>
            <li>Email: {employer['email']}</li>
            <li>Phone Number: {employer['phone']}</li>
            <li>Password: (the password you use daily)</li>
        </ul>
        <p>Thank you for partnering with PESO SmartHire!</p>
        <p>— PESO SmartHire Admin</p>
        """

        msg = Message(subject=subject, recipients=[
                      employer["email"]], html=body)
        mail.send(msg)

        cursor.close()
        conn.close()

        return jsonify({"success": True, "message": "Recruitment type change approved successfully!"})

    except Exception as e:
        print(f"[v0] Error approving recruitment type change: {e}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "message": str(e)}), 500


@admin_bp.route("/reject-recruitment-type-change/<int:employer_id>", methods=["POST"])
def reject_recruitment_type_change(employer_id):
    """
    Admin rejects employer recruitment type change
    Reverts back to old recruitment type AND restores old documents
    """
    try:
        data = request.get_json()
        reason = data.get("reason") if data else None

        conn = create_connection()
        if not conn:
            return jsonify({"success": False, "message": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM employers WHERE employer_id = %s",
            (employer_id,)
        )
        employer = cursor.fetchone()

        if not employer:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": "Employer not found"}), 404

        old_type = employer["old_recruitment_type"]

        revert_result = revert_recruitment_type_change(employer_id, conn)
        if not revert_result.get("success"):
            cursor.close()
            conn.close()
            return jsonify({"success": False, "message": f"Reversion failed: {revert_result.get('message')}"}), 500

        # Set status back to Approved, is_active to 1, clear pending flags and documents_to_reupload
        cursor.execute("""
            UPDATE employers
            SET recruitment_type = %s,
                recruitment_type_change_pending = 0,
                old_recruitment_type = NULL,
                status = 'Approved',
                is_active = 1,
                documents_to_reupload = NULL,
                approved_at = NOW()
            WHERE employer_id = %s
        """, (old_type, employer_id))

        conn.commit()

        # Send rejection email
        subject = "PESO SmartHire - Recruitment Type Change Rejected"
        reason_block = f"<p><strong>Reason:</strong> {reason}</p>" if reason else ""
        body = f"""
        <p>Dear {employer['employer_name']},</p>
        <p>This is PESO SmartHire Team.</p>
        <p>We regret to inform you that your request to change your recruitment type has been reviewed but could not be approved at this time.</p>
        {reason_block}
        <p>Your recruitment type has been reverted back to <strong>{old_type}</strong> recruitment.</p>
        <p>Your previous documents have been restored and you may continue your {old_type.lower()} recruitment activities.</p>
        <p>You may reapply for recruitment type change in the future once you meet all requirements.</p>
        <p>If you have any questions, please contact our support team.</p>
        <p>— PESO SmartHire Admin</p>
        """

        msg = Message(subject=subject, recipients=[
                      employer["email"]], html=body)
        mail.send(msg)

        cursor.close()
        conn.close()

        return jsonify({"success": True, "message": "Recruitment type change rejected successfully! Old documents restored."})

    except Exception as e:
        print(f"[v0] Error rejecting recruitment type change: {e}")
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "message": str(e)}), 500
